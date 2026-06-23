
# -*- coding: utf-8 -*-
BIST_PIE_COLOR = (213/255, 49/255, 68/255)
BOE_PIE_COLOR = (212/255, 190/255, 155/255)
BIST_COLOR = (213/255, 49/255, 68/255)
BOE_COLOR = (212/255, 190/255, 155/255)
import os
import json
import time
import re
import shutil
import tempfile
import sys
import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime, timedelta
from urllib.parse import quote
import requests
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.chart import ProjectedPieChart, Reference
from openpyxl.chart.label import DataLabelList
from PIL import Image, ImageDraw, ImageFont
try:
    import matplotlib.pyplot as plt
except ImportError:
    plt = None
try:
    import browser_cookie3
except ImportError:
    browser_cookie3 = None
try:
    from bs4 import BeautifulSoup
except ImportError:
    BeautifulSoup = None
try:
    import pythoncom
    import win32com.client as win32
except ImportError:
    pythoncom = None
    win32 = None
BASE_DIR = r"\\idmsfile1\PGM\OPR\TL Arka Ofis\BIST-ALTIN TAKİP\ALTIN RAPORU"
MAIN_DATA_DIR = os.path.join(BASE_DIR, "Veriler")
TS_DIR = os.path.join(BASE_DIR, "Raporlar")
BASE_URL = "https://uygulama.tcmb.gov.tr"
REFERER_URL = "https://uygulama.tcmb.gov.tr/MBNMALTFE/hesapBakiyeEkrani"
PROJECTION_URL = "https://uygulama.tcmb.gov.tr/DISIDIMO/aralimsle.do"
SWIFT_LIST_PATH_CANDIDATES = [
    "/osmswftse/mesaj/gelen",
    "/osmswftfe/arsiv/gelen",
    "/osmswtfe/arsiv/gelen",
]
SWIFT_DETAIL_PATH_CANDIDATES = [
    "/osmswftse/mesaj/gelen",
    "/osmswftfe/arsiv/gelen",
    "/osmswtfe/arsiv/gelen",
]
SWIFT_DETAIL_DATE_FLOOR = datetime(2025, 10, 17)
SWIFT_MESSAGE_TYPE = "MT608"
SWIFT_BOE_REF = "GOLD193"
SWIFT_ZK_BOE_REF = "GOLD212"
LEGACY_BOE_LIST_URL = f"{BASE_URL}/OSMSWFT/arsivSorgulama/getArsivGelenMesajSorguSonucList.json"
LEGACY_BOE_DETAIL_URL = f"{BASE_URL}/OSMSWFT/arsivSorgulama/getArsivMesajDetay.json"
LEGACY_APP_REFERER = f"{BASE_URL}/OSMSWFT/index.html"
LEGACY_BOE_REF_FILTER = "%gold193%"
TROY_OUNCE_TO_KG = 0.0311034768
KG_TO_TROY_OUNCE = 32.1507465
TON_TO_TROY_OUNCE = 32150.7465
WEB_ACCOUNTS = [
    "MB Altın Cinsinden Fiziki Varlıklar Hesabı",
    "MB Altın Dönüşüm Hesabı",
    "MB Hazineden Satın Alınan Altınlar",
    "MB Serbest1",
    "MB TL Karşılık Alım",
    # Sayfa 1'deki *Diğer hesabını oluşturan alt hesaplar:
    # Döviz karşılığı alınan / yurt dışından transfer edilen / konut finansmanı altın hesapları
    "MB Spot",
    "MB Döviz Karşılığı Alınan",
    "MB Döviz Karşılığı Alınan Altınlar",
    "MB Yurt Dışından Altın Transfer",
    "MB Yurt Dışından Transfer Edilen Altınlar",
    "MB Konut Finansmanı Altın Hesabı",
    "MB Genişletilmiş Konut Finansmanı Altın Hesabı",
]
TS_DISPLAY_ORDER = [
    "TL Karşılığı Alınan",
    "Altın Dönüşüm Hesabı",
    "Satım Yönlü Altın Karşılığı TL Swap Hesabı (Serbest1)",
    "Fiziki Altın Hesabı",
    "Hazineden Satın Alınan Altınlar",
    "BİST Toplam",
    "BOE",
]
TABLE_MODE_OPTIONS = [
    "Günlük",
    "Seçili Aralık - Ay Başı",
    "Seçili Aralık - Ay Sonu",
    "Son 1 Yıl - Ay Başı",
    "Son 1 Yıl - Ay Sonu",
    "Son 2 Yıl - Ay Başı",
    "Son 2 Yıl - Ay Sonu",
    "Son 5 Yıl - Ay Başı",
    "Son 5 Yıl - Ay Sonu",
]
SERIES_COLORS = ["4472C4", "ED7D31", "A5A5A5", "FFC000", "5B9BD5", "C00000", "70AD47"]
ACCOUNT_PATTERNS = {
    "TL Karşılığı Alınan": ["MB TL KARSI", "MB TL KARŞI", "MB TL KARŞILIK ALIM", "MB TL KARSILIK ALIM"],
    "Altın Dönüşüm Hesabı": ["ALTIN DONUSUM HESABI", "ALTIN DÖNÜŞÜM HESABI"],
    "Satım Yönlü Altın Karşılığı TL Swap Hesabı (Serbest1)": ["SERBEST1", "SWAP"],
    "Fiziki Altın Hesabı": ["FIZIKI VARLIK", "FİZİKİ VARLIK", "ALTIN CINSINDEN FIZIKI", "ALTIN CİNSİNDEN FİZİKİ"],
    # Sayfa 5'teki "Diğer" hesabı, Sayfa 1'deki *Diğer hesabıyla birebir aynı 3 hesaptan oluşur:
    # 1) MB PM-MB_YURTDISINDAN_ALTIN_TRANSFER
    # 2) MB PM-MB_SPOT
    # 3) MB PM-KONUT_FINANSMAN__ALTIN_HESABI
    "Diğer": [
        # --- Yurt Dışından Altın Transfer hesabının güncel ve eski adları ---
        "YURTDISINDAN ALTIN TRANSFER",
        "YURT DISINDAN ALTIN TRANSFER",
        "YURT DIŞINDAN ALTIN TRANSFER",
        "YURTDISINDAN_ALTIN_TRANSFER",
        # Eski ad: "Yurt Dışından Transfer Edilen Altınlar" (kelime sırası farklı,
        # yukarıdaki "ALTIN TRANSFER" desenleriyle eşleşmez). Zaman serisinde
        # önceki tarihli kayıtlar bu adla geldiği için ayrıca eklenmiştir.
        "YURT DIŞINDAN TRANSFER EDİLEN",
        "YURT DISINDAN TRANSFER EDILEN",
        "YURTDIŞINDAN TRANSFER EDİLEN",
        "YURTDISINDAN TRANSFER EDILEN",
        # --- MB Spot hesabının güncel ve eski adları ---
        "MB SPOT",
        "MB_SPOT",
        " SPOT",
        # Eski ad: "Döviz Karşılığı Alınan (Altınlar)". "MB Spot" hesabı geçmişte
        # bu adla raporlandığından önceki tarihli zaman serisi kayıtlarında "Diğer"
        # toplamına dahil edilebilmesi için eklenmiştir.
        "DÖVIZ KARŞILIĞI ALINAN",
        "DOVIZ KARSILIGI ALINAN",
        # --- Konut Finansmanı hesabının güncel ve eski/genişletilmiş adları ---
        "KONUT FINANSMAN",
        "KONUT FİNANSMAN",
        "KONUT_FINANSMAN",
        "KONUT_FINANSMAN__ALTIN_HESABI",
    ],
    "Hazineden Satın Alınan Altınlar": ["HAZINEDEN SATIN ALINAN", "HAZİNEDEN SATIN ALINAN"],
}
def get_output_dir():
    if getattr(sys, "frozen", False):
        return os.path.join(os.path.expanduser("~"), "Desktop")
    return TS_DIR
class GoldTimeSeriesReport:
    def __init__(self, root):
        self.root = root
        self.root.title("Altın Stok Raporu / Raporlar")
        self.root.geometry("760x500")
        self.root.eval("tk::PlaceWindow . center")
        frame = ttk.Frame(root, padding="18")
        frame.pack(expand=True, fill="both")
        ttk.Label(frame, text="Altın Stok Raporu ve Zaman Serisi", font=("Open Sans", 12, "bold")).pack(pady=(0, 12))
        form = ttk.Frame(frame)
        form.pack(fill="x", pady=(0, 12))
        self.include_time_series_var = tk.BooleanVar(value=True)
        self.ts_check = ttk.Checkbutton(
            form,
            text="Zaman serisi oluşturulsun",
            variable=self.include_time_series_var,
            command=self.toggle_time_series_inputs
        )
        self.ts_check.grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 8))
        self.start_label = ttk.Label(form, text="Başlangıç Tarihi (GG.AA.YYYY)", font=("Open Sans", 10))
        self.start_label.grid(row=1, column=0, sticky="w", padx=(0, 8), pady=4)
        self.start_entry = ttk.Entry(form, width=18)
        self.start_entry.grid(row=1, column=1, sticky="w", pady=4)
        self.end_label = ttk.Label(form, text="Bitiş Tarihi (GG.AA.YYYY)", font=("Open Sans", 10))
        self.end_label.grid(row=2, column=0, sticky="w", padx=(0, 8), pady=4)
        self.end_entry = ttk.Entry(form, width=18)
        self.end_entry.grid(row=2, column=1, sticky="w", pady=4)
        self.table_mode_label = ttk.Label(form, text="Zaman Serisi Frekansı", font=("Open Sans", 10))
        self.table_mode_label.grid(row=3, column=0, sticky="w", padx=(0, 8), pady=4)
        self.table_mode_var = tk.StringVar(value=TABLE_MODE_OPTIONS[0])
        self.table_mode_combo = ttk.Combobox(
            form,
            textvariable=self.table_mode_var,
            values=TABLE_MODE_OPTIONS,
            state="readonly",
            width=28,
        )
        self.table_mode_combo.grid(row=3, column=1, sticky="w", pady=4)
        today = datetime.today().strftime("%d.%m.%Y")
        self.end_entry.insert(0, today)
        self.start_entry.insert(0, today)
        note = (
            "Not: Önce tarayıcıdan uygulamaya giriş yapın.\n"
            "Zaman serisi seçilirse program açık tarayıcı oturumundaki cookie'leri kullanarak veriyi çeker.\n"
            "Grafik tarih adımı tarih aralığına göre otomatik belirlenir."
        )
        ttk.Label(form, text=note, foreground="#555555", justify="left", font=("Open Sans", 9)).grid(
            row=0, column=2, rowspan=4, sticky="w", padx=(20, 0)
        )
        self.toggle_time_series_inputs()
        self.btn = ttk.Button(frame, text="RAPORU OLUŞTUR", command=self.process)
        self.btn.pack(fill="x", pady=8)
        self.progress_var = tk.DoubleVar()
        self.progress = ttk.Progressbar(frame, variable=self.progress_var, maximum=100)
        self.progress.pack(fill="x", pady=10)
        self.status_var = tk.StringVar(value="Durum: Hazır")
        ttk.Label(frame, textvariable=self.status_var, font=("Open Sans", 9)).pack(anchor="w")
    def update_ui(self, percent, text):
        self.progress_var.set(percent)
        self.status_var.set(f"Durum: {text}  (%{int(percent)})")
        self.root.update_idletasks()
    def toggle_time_series_inputs(self):
        state = "normal" if self.include_time_series_var.get() else "disabled"
        for widget in [
            self.start_entry,
            self.end_entry,
            self.start_label,
            self.end_label,
            self.table_mode_label,
            self.table_mode_combo,
        ]:
            try:
                widget.configure(state=state)
            except Exception:
                pass
    @staticmethod
    def normalize(val):
        if val is None:
            return ""
        s = str(val).strip().upper()
        return s.translate(str.maketrans("İıÜüŞşĞğÇçÖö", "IIUUSSGGCCOO"))
    @staticmethod
    def to_float(val):
        if val is None:
            return 0.0
        if isinstance(val, (int, float)):
            return float(val)
        s = str(val).strip()
        if not s:
            return 0.0
        if "NET" in s.upper():
            return 0.0
        s = s.replace(" ", "")
        if "," in s and "." in s:
            last_comma = s.rfind(",")
            last_dot = s.rfind(".")
            if last_comma > last_dot:
                s = s.replace(".", "").replace(",", ".")
            else:
                s = s.replace(",", "")
        elif "," in s:
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
        cleaned = "".join(ch for ch in s if ch.isdigit() or ch in ".-")
        try:
            return float(cleaned)
        except Exception:
            return 0.0
    @staticmethod
    def add_rows(*rows):
        return [round(sum(items), 6) for items in zip(*rows)]
    @staticmethod
    def ton_to_usd_billion(ton_value, usd_price):
        usd = float(ton_value or 0.0) * TON_TO_TROY_OUNCE * float(usd_price or 0.0)
        return usd / 1_000_000_000
    @staticmethod
    def rgb_fill(r, g, b):
        return PatternFill(fill_type="solid", fgColor=f"{r:02X}{g:02X}{b:02X}")
    @staticmethod
    def today_str():
        return datetime.today().strftime("%d.%m.%Y")
    @staticmethod
    def parse_user_date(text):
        try:
            return datetime.strptime(text.strip(), "%d.%m.%Y")
        except Exception:
            raise ValueError("Tarih formatı GG.AA.YYYY olmalıdır.")
    @staticmethod
    def is_business_day(dt):
        return dt.weekday() < 5
    def get_first_business_day_of_month(self, year, month):
        dt = datetime(year, month, 1)
        while not self.is_business_day(dt):
            dt += timedelta(days=1)
        return dt
    def get_last_business_day_of_month(self, year, month):
        if month == 12:
            dt = datetime(year + 1, 1, 1) - timedelta(days=1)
        else:
            dt = datetime(year, month + 1, 1) - timedelta(days=1)
        while not self.is_business_day(dt):
            dt -= timedelta(days=1)
        return dt
    def build_all_dates(self, start_dt, end_dt):
        if start_dt > end_dt:
            raise ValueError("Başlangıç tarihi bitiş tarihinden büyük olamaz.")
        cur = end_dt
        out = []
        while cur >= start_dt:
            out.append(cur)
            cur -= timedelta(days=1)
        return out
    def get_chart_step_days(self, start_dt, end_dt):
        total_days = (end_dt - start_dt).days + 1
        if total_days <= 31:
            return 3
        elif total_days <= 62:
            return 5
        elif total_days <= 93:
            return 8
        elif total_days <= 124:
            return 10
        else:
            extra = total_days - 124
            blocks = (extra - 1) // 31 + 1
            return 10 + (blocks * 2)
    def build_chart_dates(self, start_dt, end_dt):
        step = self.get_chart_step_days(start_dt, end_dt)
        cur = end_dt
        out = []
        while cur >= start_dt:
            out.append(cur)
            cur -= timedelta(days=step)
        if all(dt.date() != end_dt.date() for dt in out):
            out.append(end_dt)
        out = sorted({dt.date(): dt for dt in out}.values(), key=lambda x: x, reverse=True)
        return out
    def build_table_dates_by_mode(self, mode, start_dt, end_dt):
        mode = (mode or "Günlük").strip()
        if mode == "Günlük":
            return self.build_all_dates(start_dt, end_dt)
        if mode.startswith("Seçili Aralık"):
            base_start = start_dt
            base_end = end_dt
        else:
            years_match = re.search(r"Son\s+(\d+)\s+Yıl", mode)
            years_back = int(years_match.group(1)) if years_match else 1
            base_end = end_dt
            try:
                base_start = end_dt.replace(year=end_dt.year - years_back)
            except ValueError:
                base_start = end_dt.replace(month=2, day=28, year=end_dt.year - years_back)
        use_month_start = "Ay Başı" in mode
        month_pairs = []
        cur = datetime(base_start.year, base_start.month, 1)
        last = datetime(base_end.year, base_end.month, 1)
        while cur <= last:
            month_pairs.append((cur.year, cur.month))
            if cur.month == 12:
                cur = datetime(cur.year + 1, 1, 1)
            else:
                cur = datetime(cur.year, cur.month + 1, 1)
        dates = []
        for year, month in month_pairs:
            dt = self.get_first_business_day_of_month(year, month) if use_month_start else self.get_last_business_day_of_month(year, month)
            if base_start <= dt <= base_end:
                dates.append(dt)
        unique_dates = []
        seen = set()
        for dt in dates:
            if dt.date() not in seen:
                unique_dates.append(dt)
                seen.add(dt.date())
        if all(dt.date() != end_dt.date() for dt in unique_dates):
            unique_dates.append(end_dt)
        unique_dates = sorted({dt.date(): dt for dt in unique_dates}.values(), key=lambda x: x, reverse=True)
        return unique_dates
    def find_latest_kasa_file(self):
        if not os.path.exists(MAIN_DATA_DIR):
            raise FileNotFoundError(f"Ağ yolu bulunamadı:\n{MAIN_DATA_DIR}")
        files = [f for f in os.listdir(MAIN_DATA_DIR) if f.startswith("Kasa Raporu_") and f.lower().endswith(".xlsx")]
        if not files:
            raise FileNotFoundError("Veriler klasöründe 'Kasa Raporu_' ile başlayan dosya bulunamadı.")
        valid_files = []
        for fname in files:
            m = re.search(r"(\d{2})[.\-\s](\d{2})[.\-\s](\d{2,4})", fname)
            if not m:
                continue
            gun, ay, yil = m.group(1), m.group(2), m.group(3)
            full_year = f"20{yil}" if len(yil) == 2 else yil
            try:
                dt = datetime(int(full_year), int(ay), int(gun))
            except ValueError:
                continue
            valid_files.append({"name": fname, "date": dt, "tag": f"{gun}{ay}{full_year[-2:]}"})
        if not valid_files:
            raise Exception("Kasa raporu dosya isimlerinden geçerli tarih okunamadı.")
        return sorted(valid_files, key=lambda x: x["date"], reverse=True)[0]
    def find_boe_file(self):
        candidates = [
            "BOE_Altınları.xlsx",
            "BOE Altınları.xlsx",
            "BOE_Altinlari.xlsx",
            "BOE Altinlari.xlsx",
        ]
        for name in candidates:
            full_path = os.path.join(MAIN_DATA_DIR, name)
            if os.path.exists(full_path):
                return full_path
        for fname in os.listdir(MAIN_DATA_DIR):
            normalized = self.normalize(fname)
            if "BOE" in normalized and "ALTIN" in normalized and fname.lower().endswith(".xlsx"):
                return os.path.join(MAIN_DATA_DIR, fname)
        raise FileNotFoundError("BOE altın dosyası bulunamadı.")
    def read_boe_data(self, boe_path):
        wb = openpyxl.load_workbook(boe_path, data_only=True)
        ws = wb.active
        usd_price = self.to_float(ws["B13"].value)
        def get_row(row_idx):
            return [self.to_float(ws.cell(row=row_idx, column=col).value) for col in [2, 3, 4, 5]]
        r_boe = get_row(5)
        r_idare = get_row(6)
        wb.close()
        r_boe = [x / 1000 for x in r_boe]
        r_idare = [x / 1000 for x in r_idare]
        # Zorunlu Karşılık Altınları -> BOE artık dosyadan okunmayacak.
        # Bu kalem yalnızca Swift GOLD212 serisinden beslenecek.
        r_zk_boe = [0.0, 0.0, 0.0, 0.0]
        return usd_price, r_boe, r_idare, r_zk_boe
    def read_boe_swap_ton(self, boe_path):
        wb = openpyxl.load_workbook(boe_path, data_only=True)
        ws = wb.active
        candidates = ["B13", "B14", "B15", "B16"]
        val = 0.0
        for cell_ref in candidates:
            tmp = self.to_float(ws[cell_ref].value)
            if tmp != 0:
                val = tmp
                break
        wb.close()
        return val
    def read_kasa_data(self, kasa_path):
        wb = openpyxl.load_workbook(kasa_path, data_only=True)
        ws = wb.active
        header_row = None
        for row in ws.iter_rows(min_row=1, max_row=25):
            normalized = [self.normalize(cell.value) for cell in row]
            if "HESAP" in normalized:
                header_row = row[0].row
                break
        if header_row is None:
            wb.close()
            raise Exception("Kasa raporunda 'HESAP' başlık satırı bulunamadı.")
        data_list = []
        for row in ws.iter_rows(min_row=header_row + 1):
            acc = str(row[1].value or "").strip()
            if not acc:
                continue
            data_list.append({
                "acc": acc,
                "acc_norm": self.normalize(acc),
                "bar": self.normalize(row[4].value),
                "lbma": self.normalize(row[5].value),
                "val": self.to_float(row[10].value)
            })
        wb.close()
        return data_list
    def calc_row(self, data_list, exact=None, contains=None, bar_type=None, exclude=None):
        total = 0.0
        exact_list = None
        if exact is not None:
            exact_list = exact if isinstance(exact, list) else [exact]
            exact_list = [self.normalize(x) for x in exact_list]
        contains_norm = self.normalize(contains) if contains else None
        exclude_norm = self.normalize(exclude) if exclude else None
        for item in data_list:
            match = False
            if exact_list is not None:
                match = item["acc_norm"] in exact_list
            elif contains_norm:
                match = contains_norm in item["acc_norm"]
                if match and exclude_norm and exclude_norm in item["acc_norm"]:
                    match = False
            if not match:
                continue
            if bar_type == 1:
                if "LARGE" in item["bar"]:
                    total += item["val"]
            elif bar_type == 2:
                if "KULCE" in item["bar"] and "ICI" in item["lbma"]:
                    total += item["val"]
            elif bar_type == 3:
                if "DISI" in item["lbma"]:
                    total += item["val"]
            else:
                total += item["val"]
        return round(total, 3)
    def get_res(self, data_list, exact=None, contains=None, exclude=None):
        large = self.calc_row(data_list, exact=exact, contains=contains, bar_type=1, exclude=exclude)
        kilo_in = self.calc_row(data_list, exact=exact, contains=contains, bar_type=2, exclude=exclude)
        kilo_out = self.calc_row(data_list, exact=exact, contains=contains, bar_type=3, exclude=exclude)
        large /= 1000
        kilo_in /= 1000
        kilo_out /= 1000
        total = round(large + kilo_in + kilo_out, 6)
        return [round(large, 6), round(kilo_in, 6), round(kilo_out, 6), total]
    def get_total_balance_ton(self, data_list, exact=None, contains=None, exclude=None):
        total = 0.0
        exact_list = None
        if exact is not None:
            exact_list = exact if isinstance(exact, list) else [exact]
            exact_list = [self.normalize(x) for x in exact_list]
        contains_norm = self.normalize(contains) if contains else None
        exclude_norm = self.normalize(exclude) if exclude else None
        for item in data_list:
            match = False
            if exact_list is not None:
                match = item["acc_norm"] in exact_list
            elif contains_norm:
                match = contains_norm in item["acc_norm"]
                if match and exclude_norm and exclude_norm in item["acc_norm"]:
                    match = False
            if match:
                total += item["val"]
        return round(total / 1000, 6)
    def create_session_from_browser(self):
        if browser_cookie3 is None:
            raise Exception(
                "browser-cookie3 paketi yüklü değil.\n"
                "Kurulum: pip --proxy proxy.tcmb.gov.tr:8080 install browser-cookie3"
            )
        session = requests.Session()
        loaded = False
        for browser_name in ["edge", "chrome", "firefox"]:
            try:
                loader = getattr(browser_cookie3, browser_name)
                # Domain filtresi bazı oturumlarda gerekli cookie'leri kaçırabiliyor.
                for kwargs in ({"domain_name": "uygulama.tcmb.gov.tr"}, {}):
                    try:
                        cj = loader(**kwargs)
                    except TypeError:
                        cj = loader()
                    if cj:
                        session.cookies.update(cj)
                        loaded = True
                if loaded:
                    break
            except Exception:
                pass
        if not loaded:
            raise Exception(
                "Tarayıcı cookie'leri okunamadı.\n"
                "Önce uygulamaya tarayıcıdan giriş yapın.\n"
                "Desteklenen tarayıcılar: Edge / Chrome / Firefox"
            )
        session.headers.update({
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:140.0) Gecko/20100101 Firefox/140.0",
            "Accept": "application/json, text/plain, */*",
            "Accept-Language": "en-US,en;q=0.5",
            "Referer": "https://uygulama.tcmb.gov.tr/osmswftfe/arsiv/gelen",
            "Sec-Fetch-Dest": "empty",
            "Sec-Fetch-Mode": "cors",
            "Sec-Fetch-Site": "same-origin",
            "Priority": "u=0",
            "Connection": "keep-alive",
        })
        return session
    def build_api_url(self, date_dt):
        date_api = date_dt.strftime("%Y-%m-%d")
        encoded_accounts = ",".join(quote(x, safe="") for x in WEB_ACCOUNTS)
        return (
            f"{BASE_URL}/mbnmaltse/hesapBakiye/hesapBakiyeListele/"
            f"Hepsi/{encoded_accounts}/Hepsi/{date_api}/{date_api}"
        )
    def fetch_day_json(self, session, date_dt):
        url = self.build_api_url(date_dt)
        resp = session.get(url, timeout=60)
        if resp.status_code != 200:
            raise Exception(f"Servis hatası: HTTP {resp.status_code}")
        return resp.json()
    def walk_json_nodes(self, obj):
        if isinstance(obj, dict):
            yield obj
            for v in obj.values():
                yield from self.walk_json_nodes(v)
        elif isinstance(obj, list):
            for x in obj:
                yield from self.walk_json_nodes(x)
    def detect_account_label(self, text_value):
        n = self.normalize(text_value)
        for target, patterns in ACCOUNT_PATTERNS.items():
            for p in patterns:
                if self.normalize(p) in n:
                    return target
        return None
    def node_to_account_value(self, node):
        if not isinstance(node, dict):
            return None, None
        possible_texts = []
        for key in ["hesapAdi", "hesap", "altHesap", "text", "name", "label"]:
            if key in node and node[key] is not None:
                possible_texts.append(str(node[key]))
        target = None
        for txt in possible_texts:
            target = self.detect_account_label(txt)
            if target:
                break
        if not target:
            return None, None
        value_keys = ["netKg", "netKG", "netkg", "net_kg", "Net Kg", "Net KG", "net", "miktar"]
        val = None
        for k in value_keys:
            if k in node:
                val = self.to_float(node.get(k))
                break
        if val is None:
            for v in node.values():
                if isinstance(v, (int, float, str)):
                    vv = self.to_float(v)
                    if vv != 0:
                        val = vv
                        break
        if val is None:
            val = 0.0
        return target, val / 1000.0
    def parse_day_response(self, json_data):
        result = {
            "TL Karşılığı Alınan": 0.0,
            "Altın Dönüşüm Hesabı": 0.0,
            "Satım Yönlü Altın Karşılığı TL Swap Hesabı (Serbest1)": 0.0,
            "Fiziki Altın Hesabı": 0.0,
            "Diğer": 0.0,
            "Hazineden Satın Alınan Altınlar": 0.0,
            "BOE": 0.0,
        }
        for node in self.walk_json_nodes(json_data):
            label, ton_val = self.node_to_account_value(node)
            if label:
                result[label] += round(ton_val, 6)
        bist_total = sum(
            result.get(label, 0.0)
            for label in [
                "TL Karşılığı Alınan",
                "Altın Dönüşüm Hesabı",
                "Satım Yönlü Altın Karşılığı TL Swap Hesabı (Serbest1)",
                "Fiziki Altın Hesabı",
                "Diğer",
                "Hazineden Satın Alınan Altınlar",
            ]
        )
        result["BİST Toplam"] = round(bist_total, 6)
        return result
    @staticmethod
    def ons_to_kg(ons_value):
        return float(ons_value or 0.0) * TROY_OUNCE_TO_KG
    @staticmethod
    def ton_to_ons(ton_value):
        return float(ton_value or 0.0) * TON_TO_TROY_OUNCE
    @staticmethod
    def ons_to_ton(ons_value):
        return float(ons_value or 0.0) / TON_TO_TROY_OUNCE
    @staticmethod
    def next_business_day(dt_value):
        nxt = dt_value + timedelta(days=1)
        while nxt.weekday() >= 5:
            nxt += timedelta(days=1)
        return nxt
    @staticmethod
    def previous_business_day(dt_value):
        prev = dt_value - timedelta(days=1)
        while prev.weekday() >= 5:
            prev -= timedelta(days=1)
        return prev
    @staticmethod
    def forward_fill_numeric_series(value_series):
        if not value_series:
            return value_series
        last_val = None
        for d in sorted(value_series.keys()):
            val = value_series.get(d)
            if val is None:
                value_series[d] = last_val
            else:
                last_val = val
        return value_series
    def build_swift_list_candidate_urls(self):
        return [f"{BASE_URL}{base_path}/page" for base_path in SWIFT_LIST_PATH_CANDIDATES]
    def build_swift_detail_candidate_urls(self, message_id):
        mids = str(message_id).strip("/")
        urls = []
        for base_path in SWIFT_DETAIL_PATH_CANDIDATES:
            urls.extend([
                f"{BASE_URL}{base_path}/{mids}",
                f"{BASE_URL}{base_path}/detay/{mids}",
                f"{BASE_URL}{base_path}/detail/{mids}",
            ])
        return urls
    def parse_json_like_response(self, resp):
        try:
            return resp.json()
        except Exception:
            pass
        text = (resp.text or "").strip()
        if not text:
            return None
        # Bazı servisler JSON ön ekleri döndürebilir.
        candidates = [text]
        stripped = text
        while stripped and stripped[0] in ")]}\n\r\t ":
            stripped = stripped[1:]
        if stripped and stripped != text:
            candidates.append(stripped)
        for candidate in candidates:
            try:
                return json.loads(candidate)
            except Exception:
                pass
        m = re.search(r'(\{.*\}|\[.*\])', text, flags=re.S)
        if m:
            candidate = m.group(1)
            try:
                return json.loads(candidate)
            except Exception:
                return None
        return None
    def fetch_swift_list_for_range(self, session, start_dt, end_dt):
        start_iso = start_dt.strftime("%Y-%m-%dT00:00:00.000")
        end_iso = end_dt.strftime("%Y-%m-%dT23:59:59.999")
        get_params = {
            "kayitTarihiBaslangic": start_iso,
            "kayitTarihiBitis": end_iso,
            "mesajTipiList": SWIFT_MESSAGE_TYPE,
            "page": 0,
            "size": 1000,
        }
        errors = []
        for url in self.build_swift_list_candidate_urls():
            try:
                resp = session.get(
                    url,
                    timeout=60,
                    params=get_params,
                    headers={
                        "Accept": "application/json, text/plain, */*",
                        "Referer": "https://uygulama.tcmb.gov.tr/osmswftfe/arsiv/gelen",
                    },
                )
                if resp.status_code != 200:
                    snippet = (resp.text or "")[:120].replace("\n", " ").replace("\r", " ")
                    errors.append(f"{url} [GET] -> HTTP {resp.status_code}: {snippet}")
                    continue
                payload = self.parse_json_like_response(resp)
                if payload is not None:
                    return payload
                snippet = (resp.text or "")[:200].replace("\n", " ").replace("\r", " ")
                errors.append(f"{url} [GET] -> JSON parse edilemedi: {snippet}")
            except Exception as exc:
                errors.append(f"{url} [GET] -> {exc}")
        raise Exception("Swift liste verisi alınamadı. " + " | ".join(errors[:4]))
    def extract_swift_list_items(self, payload):
        if isinstance(payload, list):
            flat = []
            for item in payload:
                if isinstance(item, dict):
                    flat.append(item)
                else:
                    nested = self.extract_swift_list_items(item)
                    if nested:
                        flat.extend(nested)
            return flat
        if isinstance(payload, dict):
            for key in ["content", "items", "data", "rows", "result", "results", "page", "payload"]:
                val = payload.get(key)
                nested = self.extract_swift_list_items(val)
                if nested:
                    return nested
            if any(not isinstance(v, (dict, list)) for v in payload.values()):
                return [payload]
        return []
    def extract_swift_message_id(self, item):
        if isinstance(item, (str, int)):
            return item
        if not isinstance(item, dict):
            return None
        for key in ["id", "mesajId", "messageId", "arsivMesajId", "uuid", "mesajNo", "arsivId"]:
            val = item.get(key)
            if val not in (None, ""):
                return val
        for val in item.values():
            found = self.extract_swift_message_id(val)
            if found not in (None, ""):
                return found
        return None
    def extract_swift_item_ref(self, item):
        if not isinstance(item, dict):
            return ""
        for key in ["islemReferansNo", "islemRefNo", "işlemReferansNo", "referansNo", "refNo", "islemReferans", "transactionReference"]:
            val = item.get(key)
            if val:
                return str(val)
        dump = self.normalize(json.dumps(item, ensure_ascii=False))
        m = re.search(r"GOLD\s*193", dump)
        if m:
            return "GOLD193"
        return ""
    def extract_swift_item_date(self, item):
        if not isinstance(item, dict):
            return None
        candidate_keys = ["kayitTarihi", "kayıtTarihi", "tarih", "createDate", "createdAt", "mesajTarihi", "kayitZamani", "olusturmaTarihi"]
        for key in candidate_keys:
            val = item.get(key)
            if val is None:
                continue
            if isinstance(val, datetime):
                return val
            text = str(val).strip().replace("Z", "")
            for fmt in ("%d/%m/%Y %H:%M:%S", "%d.%m.%Y %H:%M:%S", "%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S"):
                try:
                    return datetime.strptime(text[:26], fmt)
                except Exception:
                    pass
        dump = json.dumps(item, ensure_ascii=False)
        m = re.search(r"(\d{2}[/.]\d{2}[/.]\d{4})", dump)
        if m:
            raw = m.group(1).replace('/', '.').replace('-', '.')
            for fmt in ("%d.%m.%Y",):
                try:
                    return datetime.strptime(raw, fmt)
                except Exception:
                    pass
        return None
    def fetch_swift_detail_text(self, session, message_id):
        errors = []
        for url in self.build_swift_detail_candidate_urls(message_id):
            try:
                resp = session.get(url, timeout=60)
                if resp.status_code != 200:
                    errors.append(f"{url} -> HTTP {resp.status_code}")
                    continue
                payload = self.parse_json_like_response(resp)
                if payload is not None:
                    detail_text = self.extract_swift_detail_text_from_payload(payload)
                    if detail_text:
                        return detail_text
                if resp.text and ":62F:" in resp.text:
                    return resp.text
                if BeautifulSoup is not None and resp.text:
                    soup = BeautifulSoup(resp.text, "html.parser")
                    text = soup.get_text("\n", strip=True)
                    if ":62F:" in text:
                        return text
                errors.append(f"{url} -> detay metni bulunamadı")
            except Exception as exc:
                errors.append(f"{url} -> {exc}")
        raise Exception("Swift mesaj detayı alınamadı. " + " | ".join(errors[:3]))
    def extract_swift_detail_text_from_payload(self, payload):
        if isinstance(payload, str):
            return payload if ":62F:" in payload else None
        if isinstance(payload, dict):
            for key in ["mesajIcerigi", "mesajİcerigi", "mesajIcerik", "icerik", "content", "message", "rawText", "text", "mesajMetni"]:
                val = payload.get(key)
                if isinstance(val, str) and ":62F:" in val:
                    return val
            for val in payload.values():
                result = self.extract_swift_detail_text_from_payload(val)
                if result:
                    return result
        elif isinstance(payload, list):
            for item in payload:
                result = self.extract_swift_detail_text_from_payload(item)
                if result:
                    return result
        return None
    def parse_swift_62f_ounce(self, message_text):
        if not message_text:
            return None
        text = str(message_text).replace("\r", "")
        # SWIFT 62F alanında tarihten sonra gelen kıymet/para kodu 3 karakter kabul edilir.
        # Önceki sürümde kullanılan {2,5} deseni, FOZ/F0Z sonrasındaki ilk rakamları da yutup
        # 4052482,111 yerine 2482,111 veya 82,111 gibi hatalı tutarlar üretebiliyordu.
        patterns = [
            # Klasik tek satır SWIFT: :62F:C251021FOZ4052482,111
            r":62F:[A-Z]?\d{6}[A-Z0-9]{3}([0-9]+)[,\.]([0-9]+)",
            # Legacy/detail yapısı: 62F ... C220204FOZ4052482,111
            r"62F[^0-9A-Z]*[A-Z]?\d{6}[A-Z0-9]{3}([0-9]+)[,\.]([0-9]+)",
            # JSON-benzeri yapı: "kod":"62F", "icerik":"C220204FOZ4052482,111"
            r'(?:"kod"\s*:\s*"62F".{0,200}?"(?:icerik|içerik|content|value)"\s*:\s*"?[A-Z]?\d{6}[A-Z0-9]{3}([0-9]+)[,\.]([0-9]+)"?)',
        ]
        for pat in patterns:
            m = re.search(pat, text, flags=re.IGNORECASE | re.DOTALL)
            if m:
                try:
                    return float(f"{m.group(1)}.{m.group(2)}")
                except Exception:
                    pass
        # Daha gevşek ama yine 62F çevresine bağlı fallback
        m = re.search(r":62F:[^\n]*?([0-9]+[,\.][0-9]+)", text, flags=re.IGNORECASE)
        if m:
            try:
                return float(m.group(1).replace(",", "."))
            except Exception:
                pass
        m = re.search(r"62F.{0,200}?([0-9]+[,\.][0-9]+)", text, flags=re.IGNORECASE | re.DOTALL)
        if m:
            try:
                return float(m.group(1).replace(",", "."))
            except Exception:
                pass
        return None
    def fetch_gold_series_from_swift(self, session, dates_desc, message_ref, series_label):
        unique_dates = sorted({dt.date(): dt for dt in dates_desc}.values(), key=lambda x: x)
        value_series = {}
        detail_errors = []
        effective_start = max(unique_dates[0], SWIFT_DETAIL_DATE_FLOOR) if unique_dates else SWIFT_DETAIL_DATE_FLOOR
        effective_dates = [dt for dt in unique_dates if dt >= SWIFT_DETAIL_DATE_FLOOR]
        if not effective_dates:
            return value_series, detail_errors
        list_payload = self.fetch_swift_list_for_range(session, effective_start, unique_dates[-1])
        items = self.extract_swift_list_items(list_payload)
        grouped = {}
        message_ref_norm = self.normalize(message_ref)
        for item in items:
            ref_no = self.normalize(self.extract_swift_item_ref(item))
            dump_norm = self.normalize(json.dumps(item, ensure_ascii=False)) if isinstance(item, dict) else self.normalize(item)
            if message_ref_norm not in ref_no and message_ref_norm not in dump_norm:
                continue
            item_dt = self.extract_swift_item_date(item)
            if item_dt is None:
                continue
            # T-1 günü akşam gönderilen 62F mesajı, T günü raporunda kullanılır.
            # Yani 23.04 Perşembe mesajı 24.04 Cuma raporu içindir; Cuma 24.04 mesajı
            # Pazartesi 27.04 raporu içindir (hafta sonu forward-fill ile taşınır).
            target_dt = (item_dt + timedelta(days=1)).date()
            grouped.setdefault(target_dt, []).append(item)
        for dt in effective_dates:
            candidates = grouped.get(dt.date(), [])
            if not candidates:
                value_series[dt.date()] = None
                continue
            best_ounce = None
            best_msg_id = None
            candidate_errors = []
            # Yeni mantık: aynı günün tüm adaylarını dene, geçerli 62F içinden en yüksek ons değerini al.
            # Bu, yanlış alt mesaj/ara kayıt seçimini önler.
            ordered_candidates = sorted(
                candidates,
                key=lambda x: self.extract_swift_item_date(x) or dt,
                reverse=True,
            )
            for cand in ordered_candidates:
                msg_id = self.extract_swift_message_id(cand)
                ounce = None
                if msg_id not in (None, ""):
                    try:
                        detail_text = self.fetch_swift_detail_text(session, msg_id)
                        ounce = self.parse_swift_62f_ounce(detail_text)
                    except Exception as exc:
                        candidate_errors.append(f"{dt.strftime('%d.%m.%Y')} -> {series_label} msg {msg_id}: {exc}")
                if ounce is None and isinstance(cand, dict):
                    try:
                        ounce = self.parse_swift_62f_ounce(json.dumps(cand, ensure_ascii=False))
                    except Exception:
                        ounce = None
                if ounce is not None and (best_ounce is None or ounce > best_ounce):
                    best_ounce = ounce
                    best_msg_id = msg_id
            value_series[dt.date()] = round(self.ons_to_ton(best_ounce), 6) if best_ounce is not None else None
            if best_ounce is None:
                detail_errors.append(f"{dt.strftime('%d.%m.%Y')} -> {series_label} kapanış bakiyesi parse edilemedi")
                detail_errors.extend(candidate_errors[:2])
        if effective_dates and all(value_series.get(dt.date()) is None for dt in effective_dates):
            detail_errors.append(f"Swift listesinden {message_ref} için kullanılabilir {series_label} bakiyesi üretilemedi")
        return value_series, detail_errors
    def fetch_latest_gold_balance(self, session, end_dt, message_ref, series_label, lookback_days=30):
        """Rapor günü dahil geriye doğru tarayıp en son geçerli 62F mesajını bul.
        Tarih eşlemesi/forward-fill yapma — sadece kronolojik olarak en son mesajı al."""
        start_dt = end_dt - timedelta(days=lookback_days)
        list_payload = self.fetch_swift_list_for_range(session, start_dt, end_dt)
        items = self.extract_swift_list_items(list_payload)
        message_ref_norm = self.normalize(message_ref)
        # message_ref ile eşleşen tüm adayları topla, en yeni mesajdan başlayarak sırala
        candidates = []
        for item in items:
            ref_no = self.normalize(self.extract_swift_item_ref(item))
            dump_norm = self.normalize(json.dumps(item, ensure_ascii=False)) if isinstance(item, dict) else self.normalize(item)
            if message_ref_norm not in ref_no and message_ref_norm not in dump_norm:
                continue
            item_dt = self.extract_swift_item_date(item)
            if item_dt is None:
                continue
            candidates.append((item_dt, item))
        # En yeni mesajdan başlayarak parse et, ilk başarılı sonucu döndür
        candidates.sort(key=lambda x: x[0], reverse=True)
        for item_dt, cand in candidates:
            msg_id = self.extract_swift_message_id(cand)
            ounce = None
            if msg_id not in (None, ""):
                try:
                    detail_text = self.fetch_swift_detail_text(session, msg_id)
                    ounce = self.parse_swift_62f_ounce(detail_text)
                except Exception:
                    ounce = None
            if ounce is None and isinstance(cand, dict):
                try:
                    ounce = self.parse_swift_62f_ounce(json.dumps(cand, ensure_ascii=False))
                except Exception:
                    ounce = None
            if ounce is not None:
                return round(self.ons_to_ton(ounce), 6), item_dt, None
        return None, None, f"{series_label} için son {lookback_days} gün içinde geçerli 62F mesajı bulunamadı"
    def fetch_boe_series_from_swift(self, session, dates_desc):
        return self.fetch_gold_series_from_swift(session, dates_desc, SWIFT_BOE_REF, "BOE")
    def fetch_zk_boe_series_from_swift(self, session, dates_desc):
        return self.fetch_gold_series_from_swift(session, dates_desc, SWIFT_ZK_BOE_REF, "ZK BOE")
    def fetch_legacy_boe_list_for_day(self, session, date_dt):
        params = {
            "_dc": str(int(time.time() * 1000)),
            "kriterler": json.dumps({
                "arsivTarihiBaslangic": date_dt.strftime("%d/%m/%Y"),
                "arsivTarihiBitis": date_dt.strftime("%d/%m/%Y"),
                "islemRefNo": LEGACY_BOE_REF_FILTER,
            }, ensure_ascii=False),
            "page": 1,
            "start": 0,
            "limit": 25,
        }
        resp = session.get(
            LEGACY_BOE_LIST_URL,
            timeout=60,
            params=params,
            headers={
                "Accept": "*/*",
                "Referer": LEGACY_APP_REFERER,
                "X-Requested-With": "XMLHttpRequest",
                "X-TCMB-Application-Version": "2.22.1-90",
            },
        )
        if resp.status_code != 200:
            snippet = (resp.text or "")[:200].replace("\n", " ").replace("\r", " ")
            raise Exception(f"Legacy BOE liste HTTP {resp.status_code}: {snippet}")
        payload = self.parse_json_like_response(resp)
        if payload is None:
            snippet = (resp.text or "")[:200].replace("\n", " ").replace("\r", " ")
            raise Exception(f"Legacy BOE liste JSON parse edilemedi: {snippet}")
        return payload
    def extract_legacy_list_items(self, payload):
        if isinstance(payload, list):
            return payload
        if isinstance(payload, dict):
            for key in ["data", "rows", "results", "items", "content", "children", "liste", "result"]:
                val = payload.get(key)
                extracted = self.extract_legacy_list_items(val)
                if extracted:
                    return extracted
            if any(not isinstance(v, (dict, list)) for v in payload.values()):
                return [payload]
        return []
    def extract_legacy_item_mir(self, item):
        if isinstance(item, (str, int)):
            return None
        if not isinstance(item, dict):
            return None
        for key in ["mir", "MIR", "mesajMir", "messageMir"]:
            val = item.get(key)
            if val not in (None, ""):
                return str(val)
        for val in item.values():
            found = self.extract_legacy_item_mir(val)
            if found not in (None, ""):
                return found
        return None
    def extract_legacy_item_archive_date(self, item, fallback_dt=None):
        if isinstance(item, dict):
            for key in ["arsivTarihi", "arşivTarihi", "arsivTarih", "kayitTarihi", "kayıtTarihi", "tarih"]:
                val = item.get(key)
                if val in (None, ""):
                    continue
                if isinstance(val, datetime):
                    return val.date()
                text = str(val).strip().replace("Z", "")
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d.%m.%Y", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S"):
                    try:
                        return datetime.strptime(text[:19], fmt).date()
                    except Exception:
                        pass
        return fallback_dt.date() if isinstance(fallback_dt, datetime) else fallback_dt
    def fetch_legacy_boe_detail_text(self, session, arsiv_date, mir):
        params = {
            "_dc": str(int(time.time() * 1000)),
            "arsivTarihi": arsiv_date.strftime("%Y-%m-%d") if hasattr(arsiv_date, "strftime") else str(arsiv_date),
            "mir": mir,
            "page": 1,
            "start": 0,
            "limit": 25,
        }
        resp = session.get(
            LEGACY_BOE_DETAIL_URL,
            timeout=60,
            params=params,
            headers={
                "Accept": "*/*",
                "Referer": LEGACY_APP_REFERER,
                "X-Requested-With": "XMLHttpRequest",
                "X-TCMB-Application-Version": "2.22.1-90",
            },
        )
        if resp.status_code != 200:
            snippet = (resp.text or "")[:200].replace("\n", " ").replace("\r", " ")
            raise Exception(f"Legacy BOE detay HTTP {resp.status_code}: {snippet}")
        payload = self.parse_json_like_response(resp)
        if payload is not None:
            text = self.extract_swift_detail_text_from_payload(payload)
            if text:
                return text
            dumped = json.dumps(payload, ensure_ascii=False)
            if "62F" in dumped.upper():
                return dumped
        text = resp.text or ""
        if ":62F:" in text or "62F" in text.upper():
            return text
        raise Exception("Legacy BOE detay metninde 62F bulunamadı")
    def fetch_boe_series_from_legacy(self, session, dates_desc):
        unique_dates = sorted({dt.date(): dt for dt in dates_desc}.values(), key=lambda x: x)
        legacy_dates = [dt for dt in unique_dates if dt < SWIFT_DETAIL_DATE_FLOOR]
        value_series = {}
        errors = []
        for dt in legacy_dates:
            try:
                source_dt = self.previous_business_day(dt)
                payload = self.fetch_legacy_boe_list_for_day(session, source_dt)
                items = self.extract_legacy_list_items(payload)
                chosen_ounce = None
                filtered_items = []
                for item in items:
                    dump = self.normalize(json.dumps(item, ensure_ascii=False)) if isinstance(item, (dict, list)) else self.normalize(item)
                    if self.normalize("gold193") in dump:
                        filtered_items.append(item)
                candidates = filtered_items if filtered_items else items
                for item in reversed(candidates):
                    dump_text = json.dumps(item, ensure_ascii=False) if isinstance(item, (dict, list)) else str(item)
                    ounce = self.parse_swift_62f_ounce(dump_text)
                    if ounce is None:
                        mir = self.extract_legacy_item_mir(item)
                        arsiv_date = self.extract_legacy_item_archive_date(item, source_dt)
                        if mir and arsiv_date:
                            try:
                                detail_text = self.fetch_legacy_boe_detail_text(session, arsiv_date, mir)
                                ounce = self.parse_swift_62f_ounce(detail_text)
                            except Exception:
                                ounce = None
                    if ounce is not None:
                        chosen_ounce = ounce
                        break
                value_series[dt.date()] = round(self.ons_to_ton(chosen_ounce), 6) if chosen_ounce is not None else None
                if chosen_ounce is None:
                    errors.append(f"{dt.strftime('%d.%m.%Y')} -> Legacy BOE kapanış bakiyesi parse edilemedi")
            except Exception as exc:
                errors.append(f"{dt.strftime('%d.%m.%Y')} -> Legacy BOE -> {exc}")
                value_series[dt.date()] = None
        before_fill = dict(value_series)
        value_series = GoldTimeSeriesReport.forward_fill_numeric_series(value_series)
        for d in sorted(value_series.keys()):
            if before_fill.get(d) is None and value_series.get(d) is not None:
                errors = [e for e in errors if not e.startswith(f"{d.strftime('%d.%m.%Y')} -> Legacy BOE kapanış bakiyesi parse edilemedi")]
                errors.append(f"{d.strftime('%d.%m.%Y')} -> Legacy BOE için en yakın önceki geçerli bakiye kullanıldı")
        if legacy_dates and all(value_series.get(dt.date()) is None for dt in legacy_dates):
            errors.append("Legacy arşiv uygulamasından GOLD193 için kullanılabilir BOE bakiyesi üretilemedi")
        return value_series, errors
    @staticmethod
    def business_days_between(start_date, end_date):
        days = []
        cur = start_date
        while cur <= end_date:
            if cur.weekday() < 5:
                days.append(cur)
            cur += timedelta(days=1)
        return days
    def merge_boe_series(self, legacy_boe_series, swift_boe_series, ordered_dates):
        merged = {}
        all_dates = sorted({x.date() if isinstance(x, datetime) else x for x in ordered_dates})
        for dt in all_dates:
            if dt >= SWIFT_DETAIL_DATE_FLOOR.date():
                merged[dt] = swift_boe_series.get(dt)
            else:
                merged[dt] = legacy_boe_series.get(dt)
        return merged
    def forward_fill_boe_series(self, boe_series, ordered_dates):
        ordered_only = sorted({x.date() if isinstance(x, datetime) else x for x in ordered_dates})
        if not ordered_only:
            return dict(boe_series)
        filled = {}
        last_val = None
        for dt in ordered_only:
            current = boe_series.get(dt)
            if current is not None:
                last_val = current
            filled[dt] = last_val
        return filled
    def inject_boe_into_ts_results(self, ts_results, boe_series):
        out = []
        for item in ts_results:
            copied = {
                "date": item["date"],
                "date_text": item.get("date_text") or item["date"].strftime("%d.%m.%Y"),
                "values": dict(item.get("values", {})),
            }
            boe_val = boe_series.get(item["date"].date())
            copied["values"]["BOE"] = round(float(boe_val or 0.0), 6)
            copied["values"]["BİST Toplam"] = round(float(copied["values"].get("BİST Toplam", 0.0) or 0.0), 6)
            out.append(copied)
        return out
    def build_first_sheet(self, wb, usd_price, rows_data, swap_total_ton, report_date):
        ws = wb.active
        ws.title = "ALTIN STOK RAPOR"
        ws.sheet_view.showGridLines = False
        fill_row_6_27 = self.rgb_fill(84, 130, 53)
        fill_row_7 = self.rgb_fill(169, 208, 142)
        fill_row_8 = self.rgb_fill(198, 224, 180)
        fill_row_9_16 = self.rgb_fill(226, 239, 218)
        fill_row_17 = self.rgb_fill(248, 203, 173)
        fill_row_18_23 = self.rgb_fill(252, 228, 214)
        fill_row_26 = self.rgb_fill(214, 220, 228)
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        no_border = Border()
        font_header = Font(name="Open Sans", size=11, bold=True)
        font_normal = Font(name="Open Sans", size=11)
        font_bold = Font(name="Open Sans", size=11, bold=True)
        font_bold_italic = Font(name="Open Sans", size=11, bold=True, italic=True)
        font_note_bold = Font(name="Open Sans", size=11, bold=True)
        font_note = Font(name="Open Sans", size=9)
        font_date = Font(name="Open Sans", size=12, bold=True)
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
        ws.column_dimensions["A"].width = 2
        ws.column_dimensions["B"].width = 55
        ws.column_dimensions["C"].width = 14.5
        ws.column_dimensions["D"].width = 13
        ws.column_dimensions["E"].width = 13
        ws.column_dimensions["F"].width = 13
        ws.column_dimensions["G"].width = 13
        for r in range(1, 41):
            ws.row_dimensions[r].height = 15.75
        ws.row_dimensions[5].height = 18
        ws.row_dimensions[22].height = 18
        ws.row_dimensions[30].height = 30
        ws.merge_cells("B4:B5")
        ws.merge_cells("C4:D4")
        ws.merge_cells("F4:F5")
        ws.merge_cells("G4:G5")
        ws.merge_cells("B30:F30")
        ws["G2"] = report_date
        ws["G2"].font = font_date
        ws["G2"].alignment = align_center
        ws["G2"].number_format = "dd.mm.yyyy"
        ws["B4"] = "ALTINLARIN TÜRÜ"
        ws["C4"] = "LBMA İçi"
        ws["E4"] = "LBMA Dışı"
        ws["F4"] = "Toplam"
        ws["G4"] = "USD \n(Milyar USD)"
        ws["C5"] = "Large Bar"
        ws["D5"] = "Kilobar"
        ws["E5"] = "Kilobar"
        for cell in ["B4", "C4", "E4", "F4", "G4", "C5", "D5", "E5", "D4", "B5", "F5", "G5"]:
            ws[cell].border = border
            ws[cell].alignment = align_center
        for cell in ["B4", "C4", "E4", "F4", "G4", "C5", "D5", "E5"]:
            ws[cell].font = font_header
        labels = {
            6:  "Bankamız Malı Altınlar",
            7:  "İdare Merkezi ",
            8:  "BOE ",
            9:  "BİST ",
            10: "TCMB İşlemleri Kaynaklı Altınlar",
            11: "TL Karşılığı Alınan",
            12: "Altın Dönüşüm Hesabı",
            13: "Fiziki Altın Hesabı",
            14: "*Diğer",
            15: "Hazineden Satın Alınan Altınlar",
            16: "Bankamız Malı Olmayan Brüt Rezerve Dahil Altınlar",
            17: "Serbest Altın Deposu (BİST)",
            18: "Satım Yönlü Altın Karşılığı TL Swap Hesabı (Serbest1)",
            19: "Zorunlu Karşılık Altınları",
            20: "BOE",
            21: "BİST ",
            22: "Yurt Dışı Kuruluşlar Altın Depo Hesabı (BİST)",
            23: "Hazine Adına Saklanan Altınlar (BİST)",
            24: "Toplam",
            26: "* Altın Swap Stok (Yurtdışı)",
            27: "Swap Yapılan Altınlar dahil Banka Malı Altınlar",
            30: "* Diğer Hesap Döviz Karşılığı Alınan, Yurt Dışından Transfer Edilen ve Genişletilmiş Konut Finansmanı Altın Hesabından oluşmaktadır."
        }
        row_fill_map = {
            6: fill_row_6_27, 7: fill_row_7, 8: fill_row_8, 9: fill_row_9_16,
            10: fill_row_9_16, 11: fill_row_9_16, 12: fill_row_9_16, 13: fill_row_9_16,
            14: fill_row_9_16, 15: fill_row_9_16, 16: fill_row_17, 17: fill_row_18_23,
            18: fill_row_18_23, 19: fill_row_18_23, 20: fill_row_18_23, 21: fill_row_18_23,
            22: fill_row_18_23, 23: fill_row_18_23, 26: fill_row_26, 27: fill_row_6_27,
        }
        for row_no, text in labels.items():
            ws[f"B{row_no}"] = text
            if row_no == 30:
                ws[f"B{row_no}"].font = font_note
                ws[f"B{row_no}"].alignment = align_left
                ws[f"B{row_no}"].border = no_border
            else:
                if 10 <= row_no <= 15 or row_no in {17, 18, 20, 21, 22, 23}:
                    ws[f"B{row_no}"].alignment = align_right
                else:
                    ws[f"B{row_no}"].alignment = align_left
                if row_no == 10:
                    ws[f"B{row_no}"].font = font_bold_italic
                elif row_no in {6, 16, 19, 22, 23, 24, 26, 27}:
                    ws[f"B{row_no}"].font = font_note_bold if row_no in {26, 27} else font_bold
                else:
                    ws[f"B{row_no}"].font = font_normal
                ws[f"B{row_no}"].border = border
                if row_no in row_fill_map:
                    ws[f"B{row_no}"].fill = row_fill_map[row_no]
        normal_rows = list(range(6, 25))
        for row_no in normal_rows:
            for col in ["C", "D", "E", "F", "G"]:
                cell = ws[f"{col}{row_no}"]
                cell.border = border
                cell.alignment = align_right
                cell.number_format = "#,##0.00"
                if row_no in row_fill_map:
                    cell.fill = row_fill_map[row_no]
                if row_no == 10:
                    cell.font = font_bold_italic
                elif row_no in {6, 16, 19, 22, 23, 24}:
                    cell.font = font_bold
                else:
                    cell.font = font_normal
        for col in ["B", "C", "D", "E", "F", "G"]:
            ws[f"{col}24"].fill = PatternFill(fill_type=None)
        for row_no, fill_style in [(26, fill_row_26), (27, fill_row_6_27)]:
            ws[f"B{row_no}"].fill = fill_style
            ws[f"C{row_no}"].fill = fill_style
            ws[f"B{row_no}"].border = border
            ws[f"C{row_no}"].border = border
            ws[f"C{row_no}"].alignment = align_right
            ws[f"C{row_no}"].number_format = "#,##0.00"
            ws[f"C{row_no}"].font = font_note_bold
            for col in ["D", "E", "F", "G"]:
                ws[f"{col}{row_no}"].value = None
                ws[f"{col}{row_no}"].fill = PatternFill(fill_type=None)
                ws[f"{col}{row_no}"].border = no_border
                ws[f"{col}{row_no}"].font = font_normal
                ws[f"{col}{row_no}"].alignment = align_right
        def display_or_dash(value):
            if value is None:
                return "-"
            try:
                return "-" if abs(float(value)) < 1e-12 else value
            except Exception:
                return value
        def write_row(row_no, values):
            ws[f"C{row_no}"] = display_or_dash(values[0])
            ws[f"D{row_no}"] = display_or_dash(values[1])
            ws[f"E{row_no}"] = display_or_dash(values[2])
            ws[f"F{row_no}"] = display_or_dash(values[3])
            ws[f"G{row_no}"] = display_or_dash(self.ton_to_usd_billion(values[3], usd_price))
        for row_no, values in rows_data.items():
            write_row(row_no, values)
        for col in ["C", "D", "E", "F", "G"]:
            ws[f"{col}10"].font = font_bold_italic
        ws["C26"] = display_or_dash(swap_total_ton)
        ws["C27"] = display_or_dash(rows_data[6][3] + swap_total_ton)
        for col in ["B", "C", "D", "E", "F"]:
            ws[f"{col}30"].border = no_border
    def build_time_series_table_sheet(self, wb, ts_results, selected_frequency, extra_notes=None):
        ws = wb.create_sheet("ZAMAN SERİSİ_TABLO")
        ws.sheet_view.showGridLines = False
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        fill_header = self.rgb_fill(84, 130, 53)
        fill_subheader = self.rgb_fill(226, 239, 218)
        fill_total = self.rgb_fill(214, 220, 228)
        fill_grand_total = self.rgb_fill(189, 215, 238)
        font_header = Font(name="Open Sans", size=11, bold=True, color="FFFFFF")
        font_normal = Font(name="Open Sans", size=11)
        font_bold = Font(name="Open Sans", size=11, bold=True)
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = 14
        # ZAMAN SERİSİ_TABLO sayfasında "Satım Yönlü..." sütunu gösterilmiyor.
        # Sayfa 5 için "Diğer" sütunu, Sayfa 1'deki *Diğer hesabı ile aynı mantıkla
        # Fiziki Altın Hesabı ile Hazineden Satın Alınan Altınlar arasına eklenir.
        table_columns = [c for c in TS_DISPLAY_ORDER if c != "Satım Yönlü Altın Karşılığı TL Swap Hesabı (Serbest1)"]
        if "Diğer" not in table_columns:
            try:
                insert_idx = table_columns.index("Hazineden Satın Alınan Altınlar")
            except ValueError:
                insert_idx = len(table_columns)
            table_columns.insert(insert_idx, "Diğer")
        # En sağa "Toplam (BİST + BOE)" sütunu eklendi - table_columns + 1 kolon
        total_col_count = len(table_columns) + 1
        max_col_letter = openpyxl.utils.get_column_letter(2 + total_col_count)
        ws.merge_cells(f"B2:{max_col_letter}2")
        ws["B2"] = "TCMB İşlemleri Kaynaklı Altınlar (Ton)"
        for col_idx in range(2, 3 + total_col_count):
            cell = ws.cell(row=2, column=col_idx)
            cell.fill = fill_header
            cell.font = font_header
            cell.alignment = align_center
            cell.border = border
        start_col = 3
        ws["B3"] = "Tarih"
        ws["B3"].fill = fill_subheader
        ws["B3"].font = font_bold
        ws["B3"].alignment = align_center
        ws["B3"].border = border
        for idx, label in enumerate(table_columns, start=start_col):
            cell = ws.cell(row=3, column=idx, value=label)
            cell.fill = fill_subheader
            cell.font = font_bold
            cell.alignment = align_center
            cell.border = border
            ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = 18
        # En sağdaki toplam sütununun başlığı
        grand_total_col = start_col + len(table_columns)
        gt_header = ws.cell(row=3, column=grand_total_col, value="Toplam (BİST + BOE)")
        gt_header.fill = fill_subheader
        gt_header.font = font_bold
        gt_header.alignment = align_center
        gt_header.border = border
        ws.column_dimensions[openpyxl.utils.get_column_letter(grand_total_col)].width = 20
        # Sayfa 5'te BİST Toplam, solundaki 5 hesap bakiyesinin toplamı olsun.
        # Bu nedenle Satım Yönlü/Serbest1 kalemi burada BİST Toplam'a dahil edilmez.
        visible_bist_component_labels = [
            "TL Karşılığı Alınan",
            "Altın Dönüşüm Hesabı",
            "Fiziki Altın Hesabı",
            "Diğer",
            "Hazineden Satın Alınan Altınlar",
        ]
        for r_idx, item in enumerate(ts_results, start=4):
            date_cell = ws.cell(row=r_idx, column=2, value=item["date"])
            date_cell.number_format = "dd.mm.yyyy"
            date_cell.alignment = align_center
            date_cell.border = border
            date_cell.font = font_normal
            try:
                visible_bist_total = round(sum(
                    float(item["values"].get(label, 0.0) or 0.0)
                    for label in visible_bist_component_labels
                ), 6)
            except Exception:
                visible_bist_total = 0.0
            for c_idx, label in enumerate(table_columns, start=start_col):
                raw_val = item["values"].get(label, 0.0)
                display_val = visible_bist_total if label == "BİST Toplam" else raw_val
                if label == "BOE":
                    if raw_val is None or (isinstance(raw_val, (int, float)) and abs(raw_val) < 1e-9):
                        display_val = "-"
                cell = ws.cell(row=r_idx, column=c_idx, value=display_val)
                if isinstance(display_val, (int, float)):
                    cell.number_format = "#,##0.00"
                cell.alignment = align_right
                cell.border = border
                cell.font = font_bold if label in {"BİST Toplam", "BOE"} else font_normal
                if label in {"BİST Toplam", "BOE"}:
                    cell.fill = fill_total
            # En sağdaki BİST + BOE toplam hücresi
            bist_val = visible_bist_total
            boe_val = item["values"].get("BOE", 0.0) or 0.0
            try:
                grand_total_val = round(float(bist_val) + float(boe_val), 6)
            except Exception:
                grand_total_val = 0.0
            gt_cell = ws.cell(row=r_idx, column=grand_total_col, value=grand_total_val)
            gt_cell.number_format = "#,##0.00"
            gt_cell.alignment = align_right
            gt_cell.border = border
            gt_cell.font = font_bold
            gt_cell.fill = fill_grand_total
        note_lines = [f"Not: Seçilen zaman serisi frekansı = {selected_frequency}"]
        for note in (extra_notes or []):
            if note:
                note_lines.append(str(note))
        footnote_row = len(ts_results) + 4
        note_end_row = footnote_row + len(note_lines) - 1
        for offset, note_text in enumerate(note_lines):
            current_row = footnote_row + offset
            ws.merge_cells(f"B{current_row}:{max_col_letter}{current_row}")
            footnote_cell = ws[f"B{current_row}"]
            footnote_cell.value = note_text
            footnote_cell.font = Font(name="Open Sans", size=10, italic=True)
            footnote_cell.alignment = align_left
            for col_idx in range(2, 3 + total_col_count):
                ws.cell(row=current_row, column=col_idx).border = border
        last_data_row = note_end_row
        max_col = 2 + total_col_count
        for r in range(2, last_data_row + 1):
            for c in range(2, max_col + 1):
                ws.cell(row=r, column=c).border = border
        return ws
    def draw_bist_boe_distribution_image(self, wb, latest_item, output_path):
        if plt is None:
            raise Exception("matplotlib yüklü değil")
        ws_first = wb["ALTIN STOK RAPOR"]
        # Soldaki pie chart doğrudan Sheet1 F sütunundan beslensin
        boe_val = round(float(ws_first["F8"].value or 0.0), 1)
        bist_val = round(float(ws_first["F9"].value or 0.0), 1)
        # Sağdaki bar chart: BİST alt kırılımları yine Sheet1 F sütunundan
        # Nihai alt hesaplar: 11-16 satırlar
        bar_rows = [11, 12, 13, 14, 15]
        bist_parts = []
        for row_no in bar_rows:
            label = str(ws_first[f"B{row_no}"].value or "").replace("*", "").strip()
            val = float(ws_first[f"F{row_no}"].value or 0.0)
            if abs(val) > 1e-9:
                bist_parts.append((label, round(val, 1)))
        bist_parts = sorted(bist_parts, key=lambda x: x[1], reverse=True)
        pie_colors = [BOE_PIE_COLOR, BIST_PIE_COLOR]  # BOE, BİST
        bar_colors = [
            (94/255, 140/255, 198/255),
            (19/255, 119/255, 71/255),
            (157/255, 181/255, 178/255),
            (102/255, 102/255, 102/255),
            (181/255, 163/255, 182/255),
            (212/255, 190/255, 155/255),
        ]
        fig = plt.figure(figsize=(15.2, 8.2))
        fig.patch.set_facecolor("white")
        # Sol pie
        ax1 = fig.add_axes([0.03, 0.14, 0.42, 0.72])
        def _autopct(vals):
            total = sum(vals)
            def inner(pct):
                value = pct * total / 100.0
                return f"{value:,.1f}\n%{pct:.1f}"
            return inner
        ax1.pie(
            [boe_val, bist_val],
            labels=["BOE", "BİST"],
            autopct=_autopct([boe_val, bist_val]),
            pctdistance=0.63,
            labeldistance=1.06,
            startangle=90,
            colors=pie_colors,
            wedgeprops={"linewidth": 1.0, "edgecolor": "white"},
            textprops={"fontsize": 12, "fontweight": "bold"},
        )
        # Sağ bar chart
        ax2 = fig.add_axes([0.54, 0.17, 0.40, 0.64])
        if bist_parts:
            names = [n for n, _ in bist_parts]
            vals = [v for _, v in bist_parts]
            y = list(range(len(names)))
            bars = ax2.barh(
                y,
                vals,
                color=bar_colors[:len(vals)],
                edgecolor="white",
                height=0.58,
            )
            ax2.invert_yaxis()
            ax2.set_title("BİST", fontsize=14, pad=10)
            max_val = max(vals) if vals else 1
            ax2.set_xlim(-max_val * 0.80, max_val * 1.15)
            ax2.set_yticks([])
            for bar, name, val in zip(bars, names, vals):
                y_mid = bar.get_y() + bar.get_height() / 2
                ax2.annotate(
                    name,
                    xy=(0, y_mid),
                    xytext=(-max_val * 0.72, y_mid),
                    textcoords="data",
                    ha="left",
                    va="center",
                    fontsize=9,
                    arrowprops=dict(arrowstyle="-", lw=0.9, color="#666666"),
                )
                ax2.text(
                    bar.get_width() + max_val * 0.02,
                    y_mid,
                    f"{val:,.1f}",
                    va="center",
                    ha="left",
                    fontsize=10,
                    fontweight="bold",
                )
            ax2.spines["top"].set_visible(False)
            ax2.spines["right"].set_visible(False)
            ax2.spines["left"].set_visible(False)
            ax2.grid(axis="x", linestyle="--", alpha=0.18)
            ax2.tick_params(axis="x", labelsize=9)
        fig.savefig(output_path, dpi=170, bbox_inches="tight")
        plt.close(fig)
    def build_bist_boe_distribution_sheet(self, wb, latest_item, image_path):
        ws = wb.create_sheet("BİST-BOE DAĞILIM")
        ws.sheet_view.showGridLines = False
        fill_header = self.rgb_fill(84, 130, 53)
        font_header = Font(name="Open Sans", size=14, bold=True, color="FFFFFF")
        font_note = Font(name="Open Sans", size=9)
        for col in ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]:
            ws.column_dimensions[col].width = 18
        ws.merge_cells("B1:K1")
        ws["B1"] = "Bankamız Malı Altınlar - BİST ve BOE (Ton)"
        ws["B1"].font = font_header
        ws["B1"].fill = fill_header
        ws["B1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 24
        img = XLImage(image_path)
        img.width = 1260
        img.height = 710
        ws.add_image(img, "B3")
        latest_dt = latest_item.get("date")
        latest_dt_text = latest_dt.strftime("%d.%m.%Y") if isinstance(latest_dt, datetime) else str(latest_item.get("date_text", ""))
        ws["B37"] = f"Tarih: {latest_dt_text}"
        ws["B37"].font = font_note
        return ws
    def build_chart_source_sheet(self, wb, ts_results, end_dt=None):
        ws = wb.create_sheet("CHART_DATA")
        ws.sheet_state = "hidden"
        ws.sheet_view.showGridLines = False
        ws.cell(row=3, column=2, value="Tarih")
        for idx, label in enumerate(TS_DISPLAY_ORDER, start=3):
            ws.cell(row=3, column=idx, value=label)
        # Rapor tarihinin (end_dt) her zaman son satırda yer almasını garanti et
        sorted_results = sorted(ts_results, key=lambda x: x["date"])
        if end_dt is not None:
            end_date = end_dt.date() if isinstance(end_dt, datetime) else end_dt
            without_end = [x for x in sorted_results if x["date"].date() != end_date]
            end_items = [x for x in sorted_results if x["date"].date() == end_date]
            sorted_results = without_end + end_items
        for r_idx, item in enumerate(sorted_results, start=4):
            date_val = item["date"] if isinstance(item.get("date"), datetime) else datetime.strptime(str(item.get("date_text", "")), "%d.%m.%Y")
            date_cell = ws.cell(row=r_idx, column=2, value=date_val)
            date_cell.number_format = "dd.mm.yyyy"
            for c_idx, label in enumerate(TS_DISPLAY_ORDER, start=3):
                ws.cell(row=r_idx, column=c_idx, value=item["values"].get(label, 0.0))
        return ws
    @staticmethod
    def _chart_hex_to_rgb(hex_color):
        hex_color = (hex_color or "4472C4").strip().lstrip("#")
        if len(hex_color) != 6:
            return (68, 114, 196)
        return tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
    @staticmethod
    def _load_chart_font(size=16, bold=False):
        candidates = []
        if bold:
            candidates.extend([
                "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
                "C:/Windows/Fonts/arialbd.ttf",
                "C:/Windows/Fonts/calibrib.ttf",
            ])
        else:
            candidates.extend([
                "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
                "C:/Windows/Fonts/arial.ttf",
                "C:/Windows/Fonts/calibri.ttf",
            ])
        for font_path in candidates:
            try:
                return ImageFont.truetype(font_path, size)
            except Exception:
                pass
        return ImageFont.load_default()
    def draw_time_series_chart_image(self, ts_results, output_path):
        if not ts_results:
            raise ValueError("Grafik için veri bulunamadı.")
        width, height = 1600, 920
        plot_left, plot_top = 145, 55
        plot_right, plot_bottom = 1510, 610
        x_label_y = 735
        legend_top = 815
        bg_color = (255, 255, 255)
        axis_color = (60, 60, 60)
        grid_color = (222, 222, 222)
        text_color = (30, 30, 30)
        img = Image.new("RGB", (width, height), bg_color)
        draw = ImageDraw.Draw(img)
        font_axis = self._load_chart_font(22, bold=True)
        font_tick = self._load_chart_font(16, bold=False)
        font_legend = self._load_chart_font(17, bold=False)
        all_values = []
        for item in ts_results:
            for label in TS_DISPLAY_ORDER:
                all_values.append(float(item["values"].get(label, 0.0) or 0.0))
        max_val = max(all_values) if all_values else 0.0
        y_max = max(1.0, max_val * 1.12)
        y_steps = 5
        plot_width = plot_right - plot_left
        plot_height = plot_bottom - plot_top
        draw.line((plot_left, plot_top, plot_left, plot_bottom), fill=axis_color, width=2)
        draw.line((plot_left, plot_bottom, plot_right, plot_bottom), fill=axis_color, width=2)
        for i in range(y_steps + 1):
            y_val = (y_max / y_steps) * i
            y = plot_bottom - (plot_height * (i / y_steps))
            draw.line((plot_left, y, plot_right, y), fill=grid_color, width=1)
            label = f"{int(round(y_val))}"
            bbox = draw.textbbox((0, 0), label, font=font_tick)
            draw.text((plot_left - 16 - (bbox[2] - bbox[0]), y - (bbox[3] - bbox[1]) / 2), label, fill=text_color, font=font_tick)
        dates = [item.get("date_text") or item["date"].strftime("%d.%m.%Y") for item in ts_results]
        count = len(dates)
        if count == 1:
            x_positions = [plot_left + plot_width / 2]
        else:
            x_positions = [plot_left + (plot_width * idx / (count - 1)) for idx in range(count)]
        for label_index, label in enumerate(TS_DISPLAY_ORDER):
            color = self._chart_hex_to_rgb(SERIES_COLORS[label_index % len(SERIES_COLORS)])
            points = []
            for idx, item in enumerate(ts_results):
                value = float(item["values"].get(label, 0.0) or 0.0)
                y = plot_bottom - ((value / y_max) * plot_height)
                points.append((x_positions[idx], y))
            if len(points) >= 2:
                draw.line(points, fill=color, width=4)
            for x, y in points:
                draw.ellipse((x - 4, y - 4, x + 4, y + 4), fill=color, outline=color)
        for x, date_text in zip(x_positions, dates):
            bbox = draw.textbbox((0, 0), date_text, font=font_tick)
            text_w = bbox[2] - bbox[0]
            tmp = Image.new("RGBA", (text_w + 12, 36), (255, 255, 255, 0))
            tmp_draw = ImageDraw.Draw(tmp)
            tmp_draw.text((6, 4), date_text, fill=text_color, font=font_tick)
            rotated = tmp.rotate(45, expand=True)
            img.paste(rotated, (int(x - rotated.width / 2), plot_bottom + 12), rotated)
        y_title = "Ton"
        y_bbox = draw.textbbox((0, 0), y_title, font=font_axis)
        y_tmp = Image.new("RGBA", (y_bbox[2] - y_bbox[0] + 8, y_bbox[3] - y_bbox[1] + 8), (255, 255, 255, 0))
        y_tmp_draw = ImageDraw.Draw(y_tmp)
        y_tmp_draw.text((4, 4), y_title, fill=text_color, font=font_axis)
        y_rotated = y_tmp.rotate(90, expand=True)
        img.paste(y_rotated, (50, int(plot_top + plot_height / 2 - y_rotated.height / 2)), y_rotated)
        x_title = "Tarih"
        x_bbox = draw.textbbox((0, 0), x_title, font=font_axis)
        draw.text(((plot_left + plot_right - (x_bbox[2] - x_bbox[0])) / 2, x_label_y), x_title, fill=text_color, font=font_axis)
        legend_items = [(label, self._chart_hex_to_rgb(SERIES_COLORS[idx % len(SERIES_COLORS)])) for idx, label in enumerate(TS_DISPLAY_ORDER)]
        col_x = [170, 860]
        row_gap = 36
        for idx, (label, color) in enumerate(legend_items):
            col = idx % 2
            row = idx // 2
            y = legend_top + row * row_gap
            x = col_x[col]
            draw.line((x, y + 10, x + 34, y + 10), fill=color, width=4)
            draw.ellipse((x + 13, y + 3, x + 21, y + 11), fill=color, outline=color)
            draw.text((x + 48, y), label, fill=text_color, font=font_legend)
        img.save(output_path)
    def build_time_series_chart_sheet(self, wb, chart_image_path, extra_notes=None):
        ws = wb.create_sheet("ZAMAN SERİSİ_GRAFİK")
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["B"].width = 24
        ws.column_dimensions["C"].width = 24
        ws.column_dimensions["D"].width = 24
        ws.column_dimensions["E"].width = 24
        ws["B1"] = "TCMB İşlemleri Kaynaklı Altınlar"
        ws["B1"].font = Font(name="Open Sans", size=12, bold=True)
        ws["B1"].alignment = Alignment(horizontal="left", vertical="center")
        chart_img = XLImage(chart_image_path)
        chart_img.width = 1200
        chart_img.height = 690
        ws.add_image(chart_img, "B3")
        notes = [note for note in (extra_notes or []) if note]
        note_start_row = 36
        for idx, note in enumerate(notes, start=0):
            row_no = note_start_row + idx
            ws.merge_cells(start_row=row_no, start_column=2, end_row=row_no, end_column=8)
            cell = ws.cell(row=row_no, column=2, value=note)
            cell.font = Font(name="Open Sans", size=10, italic=True)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        return ws
    def draw_area_chart_image(self, ts_results, output_path):
        if plt is None:
            raise Exception("matplotlib yüklü değil")
        labels = [
            "BOE",
            "Hazineden Satın Alınan Altınlar",
            "Fiziki Altın Hesabı",
            "Satım Yönlü Altın Karşılığı TL Swap Hesabı (Serbest1)",
            "Altın Dönüşüm Hesabı",
            "TL Karşılığı Alınan",
        ]
        colors = ["#D4BE9B", "#B5A3B6", "#9DB5B2", "#137747", "#5E8CC6", "#D53144"]
        x = [item["date"] for item in ts_results]
        ys = []
        for label in labels:
            ys.append([float(item["values"].get(label, 0.0) or 0.0) for item in ts_results])
        fig, ax = plt.subplots(figsize=(13.2, 7.0))
        ax.stackplot(x, ys, labels=labels, colors=colors, alpha=0.95)
        ax.set_ylabel("Miktar (Ton)")
        ax.set_xlabel("Tarih", labelpad=12)
        ax.grid(axis="y", linestyle="--", linewidth=0.6, alpha=0.18)
        ax.grid(axis="x", visible=False)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        leg = ax.legend(loc="upper center", bbox_to_anchor=(0.5, -0.16), ncol=3, frameon=False, fontsize=9)
        fig.autofmt_xdate()
        fig.subplots_adjust(left=0.08, right=0.98, top=0.98, bottom=0.24)
        fig.savefig(output_path, dpi=170, bbox_inches="tight")
        plt.close(fig)
    def build_area_chart_sheet(self, wb, chart_data_ws, ts_results, extra_notes=None):
        from openpyxl.chart import AreaChart, Reference, Series
        from openpyxl.chart.shapes import GraphicalProperties
        from openpyxl.chart.axis import DateAxis, ChartLines
        from openpyxl.drawing.line import LineProperties
        ws = wb.create_sheet("ZAMAN SERİSİ_AREA")
        ws.sheet_view.showGridLines = False
        # CHART_DATA sayfasındaki satır aralığını bul
        data_row_start = 4
        data_row_end = data_row_start + len(ts_results) - 1
        # Sayfa başlığı: grafik başlığını grafiğin DIŞINA hücreye yaz
        ws.merge_cells("B1:N2")
        title_cell = ws["B1"]
        title_cell.value = "TCMB İşlemleri Kaynaklı Altınlar (Ton)"
        title_cell.font = Font(name="Open Sans", size=14, bold=True)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22
        ws.row_dimensions[2].height = 8
        # Alan grafik oluştur - grouping="stacked" stackplot davranışını yansıtır
        chart = AreaChart()
        chart.grouping = "stacked"
        chart.title = None  # Başlık grafik dışında, hücrede
        chart.style = 2
        chart.width = 28
        chart.height = 16
        # X eksenini DateAxis ile değiştir - tarihler sayısal/tarih olarak görünsün
        chart.x_axis = DateAxis(crossAx=100)
        chart.x_axis.number_format = "dd.mm.yy"
        chart.x_axis.majorTimeUnit = "days"
        chart.x_axis.title = "Tarih"
        chart.x_axis.title.overlay = False  # Başlık sayıların üstüne binmesin
        chart.x_axis.delete = False  # X ekseninin görünür olmasını zorla
        chart.x_axis.tickLblPos = "low"
        chart.x_axis.axPos = "b"  # Bottom - tarih ekseni grafiğin altında
        # Y eksenini açıkça görünür yap
        chart.y_axis.title = "Tutar (Ton)"
        chart.y_axis.title.overlay = False  # Başlık sayıların üstüne binmesin
        chart.y_axis.delete = False  # Y ekseninin görünür olmasını zorla
        chart.y_axis.number_format = "#,##0"
        chart.y_axis.tickLblPos = "nextTo"
        chart.y_axis.axPos = "l"  # Left - tutar ekseni grafiğin solunda
        chart.y_axis.crossAx = 500
        chart.x_axis.crossAx = chart.y_axis.axId
        # Yatay ızgara çizgilerini açık gri yap
        light_gray_line = LineProperties(solidFill="BFBFBF")
        chart.y_axis.majorGridlines = ChartLines(spPr=GraphicalProperties(ln=light_gray_line))
        # X ekseninde dikey ızgara çizgisi gösterme
        chart.x_axis.majorGridlines = None
        # TS_DISPLAY_ORDER sırası: col 3..9 → CHART_DATA sütunları
        # Alan grafikte sadece BİST alt kalemlerini ve BOE'yu çiz (BİST Toplam hariç - çift sayım)
        area_series_labels = [
            "TL Karşılığı Alınan",
            "Altın Dönüşüm Hesabı",
            "Satım Yönlü Altın Karşılığı TL Swap Hesabı (Serbest1)",
            "Fiziki Altın Hesabı",
            "Hazineden Satın Alınan Altınlar",
            "BOE",
        ]
        series_colors = ["D53144", "5E8CC6", "137747", "9DB5B2", "B5A3B6", "D4BE9B"]
        # Kategori ekseni (tarih sütunu)
        cats = Reference(chart_data_ws, min_col=2, min_row=data_row_start, max_row=data_row_end)
        for series_label, hex_color in zip(area_series_labels, series_colors):
            if series_label not in TS_DISPLAY_ORDER:
                continue
            col_idx = TS_DISPLAY_ORDER.index(series_label) + 3
            values = Reference(chart_data_ws, min_col=col_idx, min_row=data_row_start, max_row=data_row_end)
            series = Series(values, title=series_label)
            series.graphicalProperties.solidFill = hex_color
            series.graphicalProperties.line.solidFill = hex_color
            chart.series.append(series)
        chart.set_categories(cats)
        # Legend grafik altında
        chart.legend.position = "b"
        chart.legend.overlay = False
        chart.plot_area.layout = None
        ws.add_chart(chart, "B3")
        return ws
    def find_ryo_file(self):
        preferred = os.path.join(MAIN_DATA_DIR, "RYO.xlsx")
        if os.path.exists(preferred):
            return preferred
        candidates = []
        for fname in os.listdir(MAIN_DATA_DIR):
            upper = fname.upper()
            lower = fname.lower()
            if "RYO" in upper and (lower.endswith(".xlsx") or lower.endswith(".xlsm")):
                candidates.append(os.path.join(MAIN_DATA_DIR, fname))
        if not candidates:
            return None
        candidates.sort(key=lambda x: os.path.getmtime(x), reverse=True)
        return candidates[0]
    def parse_excel_date(self, val):
        if val is None or str(val).strip() == "":
            return None
        if isinstance(val, datetime):
            return val
        text = str(val).strip()
        for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d", "%d.%m.%y", "%d/%m/%y"):
            try:
                return datetime.strptime(text, fmt)
            except Exception:
                pass
        try:
            dt = openpyxl.utils.datetime.from_excel(val)
            return dt if isinstance(dt, datetime) else None
        except Exception:
            return None
    def read_ryo_projection_rows(self):
        ryo_path = self.find_ryo_file()
        if not ryo_path:
            return []
        wb = openpyxl.load_workbook(ryo_path, data_only=True)
        ws = wb.active
        try:
            control_dt = self.parse_excel_date(ws["A1000"].value)
            if control_dt is None or control_dt.date() != datetime.today().date():
                return []
            header_row_idx = None
            col_map = {}
            wanted = {
                "islem": ["ISLEM TURU", "İŞLEM TÜRÜ", "ISLEM TÜRÜ"],
                "valor": ["VALOR", "VALÖR"],
                "ons": ["ONS MIKTARI", "ONS MİKTARI"],
            }
            wanted_norm = {k: [self.normalize(x) for x in vals] for k, vals in wanted.items()}
            for r in range(1, min(ws.max_row, 50) + 1):
                row_vals = [self.normalize(ws.cell(r, c).value) for c in range(1, min(ws.max_column, 40) + 1)]
                temp_map = {}
                for idx, v in enumerate(row_vals, start=1):
                    for key, labels in wanted_norm.items():
                        if v in labels:
                            temp_map[key] = idx
                if all(k in temp_map for k in ("islem", "valor", "ons")):
                    header_row_idx = r
                    col_map = temp_map
                    break
            if header_row_idx is None:
                return []
            out = []
            blank_streak = 0
            for r in range(header_row_idx + 1, ws.max_row + 1):
                islem_val = ws.cell(r, col_map["islem"]).value
                valor_val = ws.cell(r, col_map["valor"]).value
                ons_val = ws.cell(r, col_map["ons"]).value
                if all((x is None or str(x).strip() == "") for x in (islem_val, valor_val, ons_val)):
                    blank_streak += 1
                    if blank_streak >= 20:
                        break
                    continue
                blank_streak = 0
                out.append({
                    "İşlem Türü": str(islem_val).strip() if islem_val is not None else "",
                    "Valör": valor_val,
                    "Ons Miktarı": ons_val,
                })
            return out
        finally:
            wb.close()
    def build_projection_records(self, raw_rows, report_date, start_balance_ton):
        if not raw_rows:
            return []
        rows = []
        for row in raw_rows:
            islem_turu = str(row.get("İşlem Türü", row.get("Islem Turu", ""))).strip().upper()
            valor_raw = row.get("Valör", row.get("Valor", ""))
            ton_text = row.get("Ons Miktarı", row.get("Ons Miktari", ""))
            valor_dt = self.parse_excel_date(valor_raw)
            if valor_dt is None:
                continue
            ons_val = self.to_float(ton_text)
            ton_val = self.ons_to_ton(ons_val) if ons_val != 0 else 0.0
            if valor_dt > report_date and islem_turu == "ALDIAL" and ton_val != 0:
                rows.append({"date": valor_dt, "ton": ton_val})
        if not rows:
            return []
        grouped = {}
        for item in rows:
            grouped.setdefault(item["date"], 0.0)
            grouped[item["date"]] += item["ton"]
        cumulative = None
        result = []
        for idx, dt in enumerate(sorted(grouped.keys())):
            ton = grouped[dt]
            if idx == 0:
                cumulative = start_balance_ton + ton
            else:
                cumulative += ton
            result.append({
                "date": dt,
                "ton": round(ton, 6),
                "cumulative": round(cumulative, 6),
            })
        return result
    def build_projection_sheet(self, wb, projection_records, usd_price):
        ws = wb.create_sheet("PROJEKSİYON")
        ws.sheet_view.showGridLines = False
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        fill_header = self.rgb_fill(84, 130, 53)
        fill_subheader = self.rgb_fill(226, 239, 218)
        fill_total = self.rgb_fill(214, 220, 228)
        font_header = Font(name="Open Sans", size=11, bold=True, color="FFFFFF")
        font_bold = Font(name="Open Sans", size=11, bold=True)
        font_normal = Font(name="Open Sans", size=11)
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 22
        ws.column_dimensions["E"].width = 16
        ws.merge_cells("B2:E2")
        ws["B2"] = "PROJEKSİYON"
        ws["B2"].fill = fill_header
        ws["B2"].font = font_header
        ws["B2"].alignment = align_center
        ws["B2"].border = border
        headers = ["Valör Tarihi", "Altın Swap (Yurtdışı)", "Kümülatif Bakiye-BOE (Ton)", "USD (Milyar)"]
        for idx, title in enumerate(headers, start=2):
            cell = ws.cell(row=3, column=idx, value=title)
            cell.fill = fill_subheader
            cell.font = font_bold
            cell.alignment = align_center
            cell.border = border
        total_ton = 0.0
        for r_idx, item in enumerate(projection_records, start=4):
            c_date = ws.cell(row=r_idx, column=2, value=item["date"])
            c_date.number_format = "dd.mm.yyyy"
            c_date.alignment = align_center
            c_date.border = border
            c_date.font = font_normal
            c_ton = ws.cell(row=r_idx, column=3, value=item["ton"])
            c_ton.number_format = "#,##0.00"
            c_ton.alignment = align_right
            c_ton.border = border
            c_ton.font = font_normal
            c_cum = ws.cell(row=r_idx, column=4, value=item["cumulative"])
            c_cum.number_format = "#,##0.00"
            c_cum.alignment = align_right
            c_cum.border = border
            c_cum.font = font_bold
            c_usd = ws.cell(row=r_idx, column=5, value=self.ton_to_usd_billion(item["cumulative"], usd_price))
            c_usd.number_format = "#,##0.00"
            c_usd.alignment = align_right
            c_usd.border = border
            c_usd.font = font_bold
            total_ton += float(item["ton"] or 0.0)
        total_row = len(projection_records) + 4
        ws.cell(row=total_row, column=2, value="Toplam").font = font_bold
        ws.cell(row=total_row, column=2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.cell(row=total_row, column=2).border = border
        ws.cell(row=total_row, column=2).fill = fill_total
        total_cell = ws.cell(row=total_row, column=3, value=round(total_ton, 6))
        total_cell.number_format = "#,##0.00"
        total_cell.alignment = align_right
        total_cell.border = border
        total_cell.font = font_bold
        total_cell.fill = fill_total
        for col in [4, 5]:
            last_cell = ws.cell(row=total_row, column=col, value=None)
            last_cell.border = border
            last_cell.fill = fill_total
            last_cell.alignment = align_right
            last_cell.font = font_bold
        for row in range(2, total_row + 1):
            for col in range(2, 6):
                try:
                    ws.cell(row=row, column=col).border = border
                except Exception:
                    pass
        return ws
    def export_workbook_to_pdf(self, excel_path, pdf_path):
        if win32 is None or pythoncom is None:
            raise Exception(
                "PDF üretimi için pywin32 paketi bulunamadı. "
                "Kurulum: pip install pywin32 --proxy=http://proxy.tcmb.gov.tr:8080 --trusted-host pypi.org --trusted-host files.pythonhosted.org"
            )
        pythoncom.CoInitialize()
        excel = None
        wb_com = None
        try:
            excel = win32.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            wb_com = excel.Workbooks.Open(os.path.abspath(excel_path))
            for ws in wb_com.Worksheets:
                try:
                    if ws.Visible == -1:
                        ps = ws.PageSetup
                        ps.Orientation = 2  # xlLandscape
                        ps.Zoom = False
                        ps.FitToPagesWide = 1
                        ps.FitToPagesTall = 1
                        ps.CenterHorizontally = True
                        ps.CenterVertically = False
                        ps.LeftMargin = excel.CentimetersToPoints(0.6)
                        ps.RightMargin = excel.CentimetersToPoints(0.6)
                        ps.TopMargin = excel.CentimetersToPoints(0.6)
                        ps.BottomMargin = excel.CentimetersToPoints(0.6)
                        ps.HeaderMargin = excel.CentimetersToPoints(0.3)
                        ps.FooterMargin = excel.CentimetersToPoints(0.3)
                    
                except Exception:
                    pass
            wb_com.ExportAsFixedFormat(0, os.path.abspath(pdf_path))
        finally:
            if wb_com is not None:
                try:
                    wb_com.Close(SaveChanges=False)
                except Exception:
                    pass
            if excel is not None:
                try:
                    excel.Quit()
                except Exception:
                    pass
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass
    def send_report_email(self, excel_path, pdf_path, report_date):
        if win32 is None or pythoncom is None:
            raise Exception(
                "E-posta gönderimi için pywin32 paketi bulunamadı. "
                "Kurulum: pip install pywin32 --proxy=http://proxy.tcmb.gov.tr:8080 --trusted-host pypi.org --trusted-host files.pythonhosted.org"
            )
        pythoncom.CoInitialize()
        outlook = None
        mail = None
        try:
            outlook = win32.Dispatch("Outlook.Application")
            mail = outlook.CreateItem(0)
            mail.To = "paom@tcmb.gov.tr"
            mail.Subject = f"Altın Raporu - {report_date.strftime('%d.%m.%Y')}"
            mail.Body = (
                "Hakan Bey,\n\n"
                "Bankamız malı altınlara ilişkin olarak hazırlanan Rapor ekte sunulmaktadır.\n\n"
                "Saygılarımla."
            )
            if excel_path and os.path.exists(excel_path):
                mail.Attachments.Add(os.path.abspath(excel_path))
            if pdf_path and os.path.exists(pdf_path):
                mail.Attachments.Add(os.path.abspath(pdf_path))
            mail.Send()
        finally:
            mail = None
            outlook = None
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass
    def get_time_series_results_from_web(self, session, dates_desc):
        ts_results = []
        skipped = []
        total_dates = len(dates_desc)
        for idx, dt in enumerate(dates_desc, start=1):
            progress = 40 + (idx / max(total_dates, 1)) * 35
            self.update_ui(progress, f"Web verisi alınıyor: {dt.strftime('%d.%m.%Y')}")
            try:
                raw_json = self.fetch_day_json(session, dt)
                parsed = self.parse_day_response(raw_json)
                ts_results.append({
                    "date": dt,
                    "date_text": dt.strftime("%d.%m.%Y"),
                    "values": parsed
                })
            except Exception as e:
                skipped.append(f"{dt.strftime('%d.%m.%Y')} -> {str(e)}")
        return ts_results, skipped
    def process(self):
        self.btn.config(state="disabled")
        tmp_dir = tempfile.gettempdir()
        tmp_kasa = os.path.join(tmp_dir, "kasa_report_tmp.xlsx")
        tmp_boe = os.path.join(tmp_dir, "boe_report_tmp.xlsx")
        skipped_items = []
        try:
            self.update_ui(5, "Rapor oluşturuluyor...")
            include_time_series = self.include_time_series_var.get()
            if include_time_series:
                start_dt = self.parse_user_date(self.start_entry.get())
                end_dt = self.parse_user_date(self.end_entry.get())
                if start_dt > end_dt:
                    raise Exception("Başlangıç tarihi bitiş tarihinden büyük olamaz.")
                table_mode = self.table_mode_var.get()
                table_dates = self.build_table_dates_by_mode(table_mode, start_dt, end_dt)
                if end_dt.date() not in {dt.date() for dt in table_dates}:
                    table_dates.append(end_dt)
                    table_dates = sorted(table_dates, reverse=True)
                chart_dates = self.build_chart_dates(start_dt, end_dt)
                if end_dt.date() not in {dt.date() for dt in chart_dates}:
                    chart_dates.append(end_dt)
                    chart_dates = sorted(chart_dates)
                if not table_dates:
                    raise Exception("Seçilen tarih aralığı için uygun veri tarihi bulunamadı.")
            else:
                end_dt = datetime.today()
                start_dt = end_dt
                table_dates = []
                chart_dates = []
            self.update_ui(10, "Ana rapor için en güncel kasa raporu aranıyor...")
            latest_info = self.find_latest_kasa_file()
            latest_kasa_name = latest_info["name"]
            latest_date = latest_info["date"]
            shutil.copy2(os.path.join(MAIN_DATA_DIR, latest_kasa_name), tmp_kasa)
            self.update_ui(15, "BOE dosyası okunuyor...")
            boe_source = self.find_boe_file()
            shutil.copy2(boe_source, tmp_boe)
            usd_price, r_boe, r_idare, r_zk_boe = self.read_boe_data(tmp_boe)
            self.update_ui(20, "Kasa raporu verileri okunuyor...")
            data_list = self.read_kasa_data(tmp_kasa)
            self.update_ui(28, "Ana rapor hesaplamaları yapılıyor...")
            res_hzn = self.get_res(data_list, exact="MB PM-MB_HAZINEDEN_SATIN_ALINAN_ALTINLAR")
            res_tl = self.get_res(data_list, exact="MB PM-MB_TLKARSILIKALIM")
            res_don = self.get_res(data_list, exact="MB PM-MB_ALTIN_DONUSUM_HESABI")
            res_swap_s1 = self.get_res(data_list, contains="MB SERBEST1")
            res_fiziki = self.get_res(data_list, exact="MB PM-ALTIN_CINSINDEN_FIZIKI_VARLIKLAR")
            res_digery = self.get_res(
                data_list,
                exact=[
                    "MB PM-MB_YURTDISINDAN_ALTIN_TRANSFER",
                    "MB PM-MB_SPOT",
                    "MB PM-KONUT_FINANSMAN__ALTIN_HESABI"
                ]
            )
            res_tcmb_islem = self.add_rows(res_tl, res_don, res_fiziki, res_digery)
            res_bist = self.add_rows(res_tl, res_don, res_fiziki, res_digery, res_hzn)
            res_banka = self.add_rows(r_idare, r_boe, res_bist)
            res_serbest_base = self.get_res(data_list, contains="MB SERBEST", exclude="MB SERBEST1")
            res_serbest = self.add_rows(res_serbest_base, res_swap_s1)
            res_zk_bist = self.get_res(data_list, contains="_ZKH")
            res_zk_top = self.add_rows(res_zk_bist, r_zk_boe)
            res_yd_kur = self.get_res(data_list, contains="MB PM-LIBYA_MERKEZ BANKASI_HESABI")
            res_saklama = self.get_res(data_list, exact="MB PM-HB_HAZINE_ADINA_SAKLAMA")
            res_not_bank = self.add_rows(res_serbest_base, res_swap_s1, res_zk_top, res_yd_kur, res_saklama)
            res_total = self.add_rows(res_banka, res_not_bank)
            projection_records = []
            projection_raw_rows = []
            try:
                self.update_ui(35, "RYO projeksiyon verisi okunuyor...")
                projection_raw_rows = self.read_ryo_projection_rows()
                projection_records = self.build_projection_records(projection_raw_rows, end_dt, r_boe[0])
            except Exception as proj_exc:
                skipped_items.append(f"Projeksiyon verisi alınamadı -> {str(proj_exc)}")
            swap_total_ton = 0.0
            if projection_records:
                swap_total_ton = round(sum(float(item.get("ton", 0.0) or 0.0) for item in projection_records), 6)
            rows_data = {
                6: res_banka, 7: r_idare, 8: r_boe, 9: res_bist, 10: res_tcmb_islem, 11: res_tl,
                12: res_don, 13: res_fiziki, 14: res_digery, 15: res_hzn,
                16: res_not_bank, 17: res_serbest_base, 18: res_swap_s1, 19: res_zk_top, 20: r_zk_boe,
                21: res_zk_bist, 22: res_yd_kur, 23: res_saklama, 24: res_total,
            }
            self.update_ui(40, "Workbook oluşturuluyor...")
            wb = Workbook()
            self.build_first_sheet(
                wb=wb,
                usd_price=usd_price,
                rows_data=rows_data,
                swap_total_ton=swap_total_ton,
                report_date=end_dt
            )
            ws_table = None
            ws_chart_data = None
            ws_distribution = None
            if include_time_series:
                self.update_ui(45, "Tarayıcı oturumu okunuyor...")
                session = self.create_session_from_browser()
                combined_dates = sorted(
                    {dt.date(): dt for dt in (table_dates + chart_dates)}.values(),
                    key=lambda x: x,
                    reverse=True,
                )
                all_results, skipped_web = self.get_time_series_results_from_web(session, combined_dates)
                skipped_items.extend(skipped_web)
                if not all_results:
                    raise Exception("Web uygulamasından işlenebilen zaman serisi verisi alınamadı.")
                # Kritik düzeltme:
                # BOE carry-forward seçilmiş seri tarihleri arasında değil, tüm iş günleri üzerinden yapılmalı.
                # Böylece örn. 14.04.2026 boşsa 31.03.2026'ya değil 13.04.2026'ya gider.
                #
                # Sayfa 1'deki BOE/ZK BOE değerleri "bir önceki iş gününün KAPANIŞ bakiyesi" olmalı.
                # Swift mantığı: bir önceki iş günü (örn Cuma) gönderdiği mesaj o günün kapanış
                # bakiyesidir ve fetch_gold_series_from_swift içinde target_dt = item_dt + 1 gün
                # olarak gruplanır. Pazartesi raporunda Cuma mesajı target_dt = Cumartesi key'ine
                # düşer; forward-fill ile Pazartesi (end_dt) key'ine taşınır. Bu nedenle Sayfa 1'de
                # kullanılması gereken key end_dt.date()'dir; prev_business_dt.date() kullanmak
                # bir önceki iş gününün kapanışını (Perşembe verisini) verir.
                prev_business_dt = self.previous_business_day(end_dt)
                # Tatil/bayram emniyeti: önceki iş günü resmi tatilse o gün için Swift mesajı
                # gelmez. Forward-fill için bir önceki geçerli mesajı bulabilmek adına lookup
                # aralığını 14 takvim günü daha geriye uzatıyoruz. Böylece 9 günlük bayram
                # tatillerinde bile bir önceki geçerli BOE bakiyesine ulaşabiliriz.
                BOE_LOOKBACK_BUFFER_DAYS = 14
                boe_lookup_dates = []
                if combined_dates:
                    raw_start = min(min(dt.date() for dt in combined_dates), prev_business_dt.date())
                    lookup_start = raw_start - timedelta(days=BOE_LOOKBACK_BUFFER_DAYS)
                    lookup_end = max(max(dt.date() for dt in combined_dates), end_dt.date())
                    boe_lookup_dates = [datetime.combine(d, datetime.min.time()) for d in self.business_days_between(lookup_start, lookup_end)]
                else:
                    buffer_start = (prev_business_dt - timedelta(days=BOE_LOOKBACK_BUFFER_DAYS)).date()
                    buffer_end = end_dt.date()
                    boe_lookup_dates = [datetime.combine(d, datetime.min.time()) for d in self.business_days_between(buffer_start, buffer_end)]
                self.update_ui(70, "BOE verileri alınıyor...")
                legacy_boe_series, skipped_legacy_boe = self.fetch_boe_series_from_legacy(session, boe_lookup_dates)
                skipped_items.extend(skipped_legacy_boe)
                swift_boe_series, skipped_swift = self.fetch_boe_series_from_swift(session, boe_lookup_dates)
                skipped_items.extend(skipped_swift)
                boe_series_full = self.merge_boe_series(legacy_boe_series, swift_boe_series, boe_lookup_dates)
                boe_series_full = self.forward_fill_boe_series(boe_series_full, boe_lookup_dates)
                # Seçilen seri tarihlerine projekte et
                boe_series = {dt.date(): boe_series_full.get(dt.date()) for dt in combined_dates}
                swift_zk_boe_series, skipped_swift_zk = self.fetch_zk_boe_series_from_swift(session, boe_lookup_dates)
                skipped_items.extend(skipped_swift_zk)
                swift_zk_boe_series_full = self.forward_fill_boe_series(swift_zk_boe_series, boe_lookup_dates)
                swift_zk_boe_series = {dt.date(): swift_zk_boe_series_full.get(dt.date()) for dt in combined_dates}
                # SAYFA 1 BOE / ZK BOE: Basit mantık — rapor gününe kadar olan en son
                # geçerli 62F mesajını al. Tarih eşlemesi/forward-fill yok.
                latest_boe_ton, latest_boe_dt, boe_err = self.fetch_latest_gold_balance(
                    session, end_dt, SWIFT_BOE_REF, "BOE"
                )
                if latest_boe_ton is not None:
                    r_boe = [round(latest_boe_ton, 6), 0.0, 0.0, round(latest_boe_ton, 6)]
                elif boe_err:
                    skipped_items.append(boe_err)
                latest_zk_ton, latest_zk_dt, zk_err = self.fetch_latest_gold_balance(
                    session, end_dt, SWIFT_ZK_BOE_REF, "ZK BOE"
                )
                if latest_zk_ton is not None:
                    r_zk_boe = [round(latest_zk_ton, 6), 0.0, 0.0, round(latest_zk_ton, 6)]
                elif zk_err:
                    skipped_items.append(zk_err)
                # === DEBUG: Sayfa 1 BOE seçimini diske yaz ===
                try:
                    import os as _os_dbg
                    debug_path = _os_dbg.path.join(_os_dbg.path.expanduser("~"), "Desktop", "boe_debug.txt")
                    with open(debug_path, "w", encoding="utf-8") as _df:
                        _df.write(f"=== SAYFA 1 BOE DEBUG ===\n")
                        _df.write(f"Rapor günü: {end_dt.strftime('%d.%m.%Y %A')}\n\n")
                        _df.write(f"GOLD193 (BOE) - en son mesaj:\n")
                        if latest_boe_dt is not None:
                            _df.write(f"  Mesaj tarihi: {latest_boe_dt.strftime('%d.%m.%Y %H:%M:%S %A')}\n")
                        _df.write(f"  Tonaj:        {latest_boe_ton}\n")
                        if boe_err: _df.write(f"  HATA: {boe_err}\n")
                        _df.write(f"\nGOLD212 (ZK BOE) - en son mesaj:\n")
                        if latest_zk_dt is not None:
                            _df.write(f"  Mesaj tarihi: {latest_zk_dt.strftime('%d.%m.%Y %H:%M:%S %A')}\n")
                        _df.write(f"  Tonaj:        {latest_zk_ton}\n")
                        if zk_err: _df.write(f"  HATA: {zk_err}\n")
                        _df.write(f"\nSayfa 1'e yazılan:\n")
                        _df.write(f"  r_boe:    {r_boe}\n")
                        _df.write(f"  r_zk_boe: {r_zk_boe}\n")
                except Exception as _dbg_exc:
                    skipped_items.append(f"Debug yazılamadı -> {_dbg_exc}")
                # === DEBUG SONU ===
                res_zk_top = self.add_rows(res_zk_bist, r_zk_boe)
                res_banka = self.add_rows(r_idare, r_boe, res_bist)
                res_not_bank = self.add_rows(res_serbest, res_zk_top, res_yd_kur, res_saklama)
                res_total = self.add_rows(res_banka, res_not_bank)
                rows_data = {
                    6: res_banka, 7: r_idare, 8: r_boe, 9: res_bist, 10: res_tcmb_islem, 11: res_tl,
                    12: res_don, 13: res_fiziki, 14: res_digery, 15: res_hzn,
                    16: res_not_bank, 17: res_serbest_base, 18: res_swap_s1, 19: res_zk_top, 20: r_zk_boe,
                    21: res_zk_bist, 22: res_yd_kur, 23: res_saklama, 24: res_total,
                }
                self.build_first_sheet(
                    wb=wb,
                    usd_price=usd_price,
                    rows_data=rows_data,
                    swap_total_ton=swap_total_ton,
                    report_date=end_dt
                )
                if projection_raw_rows:
                    projection_records = self.build_projection_records(projection_raw_rows, end_dt, r_boe[0])
                    swap_total_ton = round(sum(float(item.get("ton", 0.0) or 0.0) for item in projection_records), 6) if projection_records else 0.0
                all_results = self.inject_boe_into_ts_results(all_results, boe_series)
                # Rapor günü (end_dt) için Sayfa 5'teki BOE değeri Sayfa 1 ile aynı olsun:
                # ikisi de "rapor gününe kadarki en son geçerli GOLD193 mesajı" mantığı kullansın.
                if latest_boe_ton is not None:
                    end_dt_date = end_dt.date()
                    for _item in all_results:
                        if _item["date"].date() == end_dt_date:
                            _item["values"]["BOE"] = round(float(latest_boe_ton), 6)
                results_by_date = {item["date"].date(): item for item in all_results}
                # Sayfa 5 rapor günü satırı Sayfa 1 ile birebir tutarlı olsun.
                # Tarihsel satırlar web/zaman serisi kaynağından gelmeye devam eder;
                # yalnızca en güncel rapor günü satırı Kasa Raporu kaynaklı Sayfa 1 değerleriyle override edilir.
                end_dt_key = end_dt.date()
                if end_dt_key in results_by_date:
                    latest_values = results_by_date[end_dt_key]["values"]
                    sheet1_tl = round(float((res_tl[3] if len(res_tl) > 3 else 0.0) or 0.0), 6)
                    sheet1_don = round(float((res_don[3] if len(res_don) > 3 else 0.0) or 0.0), 6)
                    sheet1_fiziki = round(float((res_fiziki[3] if len(res_fiziki) > 3 else 0.0) or 0.0), 6)
                    sheet1_diger = round(float((res_digery[3] if len(res_digery) > 3 else 0.0) or 0.0), 6)
                    sheet1_hzn = round(float((res_hzn[3] if len(res_hzn) > 3 else 0.0) or 0.0), 6)
                    sheet1_boe = round(float((r_boe[3] if len(r_boe) > 3 else r_boe[0]) or 0.0), 6)
                    latest_values["TL Karşılığı Alınan"] = sheet1_tl
                    latest_values["Altın Dönüşüm Hesabı"] = sheet1_don
                    latest_values["Fiziki Altın Hesabı"] = sheet1_fiziki
                    latest_values["Diğer"] = sheet1_diger
                    latest_values["Hazineden Satın Alınan Altınlar"] = sheet1_hzn
                    latest_values["BOE"] = sheet1_boe
                    latest_values["BİST Toplam"] = round(
                        sheet1_tl + sheet1_don + sheet1_fiziki + sheet1_diger + sheet1_hzn,
                        6
                    )
                    latest_values["Toplam"] = latest_values["BİST Toplam"]
                table_results = [results_by_date[dt.date()] for dt in table_dates if dt.date() in results_by_date]
                # Sayfa 1 - Sayfa 5 BOE tutarlılığı:
                # Rapor günü satırı Sayfa 5'te varsa, BOE değeri Sayfa 1 F8 ile aynı olmalı.
                sheet1_boe_total = round(float((r_boe[3] if len(r_boe) > 3 else r_boe[0]) or 0.0), 6)
                for _item in table_results:
                    if _item["date"].date() == end_dt.date():
                        current_boe = round(float(_item["values"].get("BOE", 0.0) or 0.0), 6)
                        if abs(current_boe - sheet1_boe_total) > 0.000001:
                            _item["values"]["BOE"] = sheet1_boe_total
                            skipped_items.append(
                                f"BOE tutarlılık düzeltmesi: Sayfa 5 rapor günü BOE değeri Sayfa 1 BOE değeriyle eşitlendi ({sheet1_boe_total:.2f} ton)."
                            )
                # Sayfa 5 "Diğer" tutarlılığı:
                # Rapor günü satırında Sayfa 5'teki "Diğer" değeri, Sayfa 1'deki *Diğer hesabı ile
                # birebir aynı 3 hesabın toplamı olsun:
                #   1) MB PM-MB_YURTDISINDAN_ALTIN_TRANSFER
                #   2) MB PM-MB_SPOT
                #   3) MB PM-KONUT_FINANSMAN__ALTIN_HESABI
                sheet1_diger_total = round(float((res_digery[3] if len(res_digery) > 3 else 0.0) or 0.0), 6)
                visible_bist_component_labels = [
                    "TL Karşılığı Alınan",
                    "Altın Dönüşüm Hesabı",
                    "Fiziki Altın Hesabı",
                    "Diğer",
                    "Hazineden Satın Alınan Altınlar",
                ]
                for _item in table_results:
                    if _item["date"].date() == end_dt.date():
                        old_diger = round(float(_item["values"].get("Diğer", 0.0) or 0.0), 6)
                        if abs(old_diger - sheet1_diger_total) > 0.000001:
                            _item["values"]["Diğer"] = sheet1_diger_total
                            _item["values"]["BİST Toplam"] = round(
                                sum(float(_item["values"].get(label, 0.0) or 0.0) for label in visible_bist_component_labels),
                                6
                            )
                            skipped_items.append(
                                f"Diğer tutarlılık düzeltmesi: Sayfa 5 rapor günü Diğer değeri Sayfa 1 *Diğer değeriyle eşitlendi ({sheet1_diger_total:.2f} ton)."
                            )
                # Grafik, tablo frekansından bağımsız kendi tarih setiyle beslensin.
                chart_results = [results_by_date[dt.date()] for dt in chart_dates if dt.date() in results_by_date]
                if not table_results:
                    raise Exception("Zaman serisi tablosu için uygun veri alınamadı.")
                # Grafik seti beklenenden dar kaldıysa, tüm çekilen sonuçlardan kronolojik seri kur.
                if len(chart_results) < 2:
                    chart_results = sorted(all_results, key=lambda x: x["date"])
                # Yine de boşsa tabloyu fallback olarak kullan.
                if not chart_results:
                    chart_results = sorted(table_results, key=lambda x: x["date"])
                table_results = sorted(table_results, key=lambda x: x["date"])
                chart_results = sorted(chart_results, key=lambda x: x["date"])
                # Aynı tarihleri tekilleştir
                seen_dates = set()
                cleaned_chart_results = []
                for item in chart_results:
                    item_date = item["date"].date()
                    if item_date not in seen_dates:
                        cleaned_chart_results.append(item)
                        seen_dates.add(item_date)
                chart_results = cleaned_chart_results
                ts_notes = []
                if start_dt < SWIFT_DETAIL_DATE_FLOOR:
                    ts_notes.append(
                        f"BOE verileri {SWIFT_DETAIL_DATE_FLOOR.strftime('%d.%m.%Y')} tarihinden itibaren Swift uygulamasından, daha eski tarihler için eski arşiv uygulamasından alınmaktadır."
                    )
                ts_notes.append("BOE verisi bulunmayan iş günlerinde önceki en güncel BOE değeri kullanılmıştır.")
                self.update_ui(80, "Zaman serisi tablosu oluşturuluyor...")
                ws_table = self.build_time_series_table_sheet(wb, table_results, table_mode, extra_notes=ts_notes)
                ws_chart_data = self.build_chart_source_sheet(wb, chart_results, end_dt=end_dt)
                latest_distribution_item = (sorted(table_results, key=lambda x: x.get('date') or datetime.min, reverse=True)[0] if table_results else None)
                ws_distribution = None
                if latest_distribution_item is not None:
                    try:
                        distribution_image_path = os.path.join(tmp_dir, "bist_boe_distribution.png")
                        self.draw_bist_boe_distribution_image(wb, latest_distribution_item, distribution_image_path)
                        ws_distribution = self.build_bist_boe_distribution_sheet(wb, latest_distribution_item, distribution_image_path)
                    except Exception as dist_exc:
                        skipped_items.append(f"BİST-BOE grafik sayfası üretilemedi -> {str(dist_exc)}")
                self.update_ui(90, "Alan grafik oluşturuluyor...")
                ws_area = self.build_area_chart_sheet(wb, ws_chart_data, chart_results, extra_notes=ts_notes)
            ws_projection = None
            if projection_records:
                ws_projection = self.build_projection_sheet(wb, projection_records, usd_price)
            ws_first = wb["ALTIN STOK RAPOR"]
            ordered_sheets = [ws_first]
            if ws_projection is not None:
                ordered_sheets.append(ws_projection)
            if include_time_series:
                if ws_distribution is not None:
                    ordered_sheets.append(ws_distribution)
                ordered_sheets.extend([ws_area, ws_table, ws_chart_data])
            wb._sheets = ordered_sheets
            self.update_ui(95, "Dosya kaydediliyor...")
            output_dir = get_output_dir()
            if not os.path.exists(output_dir):
                os.makedirs(output_dir, exist_ok=True)
            file_date = self.today_str()
            base_name = f"ALTIN STOK RAPOR_{file_date}"
            final_path = os.path.join(output_dir, f"{base_name}.xlsx")
            counter = 1
            while os.path.exists(final_path):
                final_path = os.path.join(output_dir, f"{base_name}({counter}).xlsx")
                counter += 1
            wb.save(final_path)
            wb.close()
            pdf_path = os.path.splitext(final_path)[0] + ".pdf"
            try:
                self.update_ui(98, "PDF oluşturuluyor...")
                self.export_workbook_to_pdf(final_path, pdf_path)
            except Exception as pdf_exc:
                skipped_items.append(f"PDF üretilemedi -> {str(pdf_exc)}")
            try:
                self.update_ui(99, "E-posta gönderiliyor...")
                self.send_report_email(final_path, pdf_path, end_dt)
            except Exception as mail_exc:
                skipped_items.append(f"E-posta gönderilemedi -> {str(mail_exc)}")
            self.update_ui(100, "Rapor oluşturuldu ✔")
            if skipped_items:
                msg = (
                    "Rapor başarıyla oluşturuldu.\n\n"
                    f"Excel Dosyası:\n{os.path.basename(final_path)}\n\n"
                    f"PDF Dosyası:\n{os.path.basename(pdf_path)}\n\n"
                    "İşlenemeyen öğeler:\n- " + "\n- ".join(skipped_items[:20])
                )
                if len(skipped_items) > 20:
                    msg += f"\n... ve {len(skipped_items) - 20} öğe daha."
                messagebox.showwarning("Başarılı (Uyarılı)", msg)
            else:
                messagebox.showinfo("Başarılı", f"Rapor başarıyla oluşturuldu.\n\nExcel Dosyası:\n{os.path.basename(final_path)}\n\nPDF Dosyası:\n{os.path.basename(pdf_path)}")
        except Exception as e:
            self.update_ui(0, "Hata oluştu")
            messagebox.showerror("Hata", str(e))
        finally:
            self.btn.config(state="normal")
            for f in [tmp_kasa, tmp_boe, os.path.join(tmp_dir, "altin_stok_chart.png")]:
                try:
                    if os.path.exists(f):
                        os.remove(f)
                except Exception:
                    pass
if __name__ == "__main__":
    root = tk.Tk()
    app = GoldTimeSeriesReport(root)
    root.mainloop()
